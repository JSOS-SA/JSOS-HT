"""منشئ التقارير — خيار 11 في القائمة الرئيسية.

يقرأ ملف إكسل، يعرض الأعمدة، يختار المستخدم البيانات ونوع الرسم
ثم يولّد PDF بالرسوم المختارة بدون تحديد عدد.
"""

__all__ = ["run"]

import contextlib
import os
import sys
import tempfile
import warnings
from datetime import datetime
from datetime import time as dt_time
from pathlib import Path

import openpyxl
import pandas as pd

from common import (
    BOLD,
    CYAN,
    DIM,
    GREEN,
    RESET,
    ask,
    ask_file,
    ask_file_dialog,
    pause,
    print_err,
    print_header,
    print_ok,
    print_title,
    print_warn,
)
from report_charts import (
    auto_summaries,
    generate_chart,
    inject_dispatch_columns,
    normalize_code_columns,
)
from report_pdf import build_pdf

# === أوامر التنقل بين الخطوات ===
NAV_BACK = "back"
NAV_HOME = "home"
NAV_RETRY = "retry"
TOTAL_STEPS = 4


def _check_nav(user_input: str) -> str | None:
    """فحص إدخال المستخدم لأوامر التنقل — يعيد نوع الأمر أو None."""
    cleaned = user_input.strip()
    if cleaned == "0":
        return NAV_BACK
    if cleaned in ("ر", "r"):
        return NAV_HOME
    if cleaned in ("ع", "a"):
        return NAV_RETRY
    return None


def _show_step_header(step_num: int, title: str) -> None:
    """عرض رأس الخطوة مع شريط التقدم وخيارات التنقل."""
    filled = "█" * step_num
    empty = "░" * (TOTAL_STEPS - step_num)
    print(f"\n  {BOLD}[{filled}{empty}]{RESET} {step_num}/{TOTAL_STEPS}")
    print_header(title)
    print(f"  {DIM}0{RESET} رجوع خطوة")
    print(f"  {DIM}ر{RESET} القائمة الرئيسية")
    print(f"  {DIM}ع{RESET} إعادة الخطوة")
    print()


def _read_record_info(file_path: str) -> dict:
    """قراءة التاريخ وبيانات النوبات من ورقة Record — تُستخدم في هيدر PDF."""
    info = {"date": None, "shifts": []}
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        record_sheet = None
        for sn in wb.sheetnames:
            if sn.lower().strip() == "record":
                record_sheet = wb[sn]
                break
        if record_sheet is None:
            wb.close()
            return info

        rows_data = []
        for row in record_sheet.iter_rows(max_row=6, values_only=False):
            rows_data.append(row)

        if len(rows_data) >= 6:
            date_cell = rows_data[5][0].value
            if date_cell is not None:
                if hasattr(date_cell, "strftime"):
                    info["date"] = date_cell.strftime("%d/%m/%Y")
                else:
                    info["date"] = str(date_cell).split(" ")[0]

        shift_names = {"A": "الصباحية", "B": "المسائية", "C": "الليلية"}
        for idx in range(2, 5):
            if idx < len(rows_data):
                row = rows_data[idx]
                letter = str(row[13].value).strip() if row[13].value else None
                buses = row[14].value if len(row) > 14 else 0
                pax = row[15].value if len(row) > 15 else 0
                if letter in shift_names:
                    info["shifts"].append(
                        {
                            "letter": letter,
                            "name": shift_names[letter],
                            "buses": int(buses) if buses else 0,
                            "pax": int(pax) if pax else 0,
                        },
                    )

        wb.close()
    except Exception:
        pass
    return info


def _detect_header_row(
    filepath: str,
    sheet_name: str,
    max_scan: int = 30,
) -> tuple[int, int]:
    """كشف تلقائي لصف العناوين وصف بداية البيانات."""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)

    ws = wb[sheet_name]
    total_cols = ws.max_column or 1

    best_header = 0
    best_score = 0.0

    for row_idx in range(1, min(max_scan + 1, (ws.max_row or 1) + 1)):
        cells = next(
            iter(
                ws.iter_rows(
                    min_row=row_idx,
                    max_row=row_idx,
                    max_col=total_cols,
                    values_only=False,
                )
            )
        )

        non_empty = 0
        str_count = 0
        bold_count = 0

        for cell in cells:
            if cell.value is not None:
                non_empty += 1
                if isinstance(cell.value, str):
                    str_count += 1
            if cell.font and cell.font.bold:
                bold_count += 1

        if non_empty == 0:
            continue

        str_ratio = str_count / non_empty
        coverage = non_empty / total_cols
        bold_bonus = 0.1 if bold_count > total_cols * 0.3 else 0
        score = (str_ratio * 0.5) + (coverage * 0.4) + bold_bonus

        if str_ratio >= 0.8 and coverage >= 0.4 and score > best_score:
            best_score = score
            best_header = row_idx - 1

    data_start = best_header + 1
    for row_idx in range(best_header + 2, min(best_header + 10, (ws.max_row or 1) + 1)):
        cells = next(
            iter(
                ws.iter_rows(
                    min_row=row_idx,
                    max_row=row_idx,
                    max_col=total_cols,
                    values_only=True,
                )
            )
        )

        non_empty = sum(1 for v in cells if v is not None)
        has_nums = any(isinstance(v, (int, float)) for v in cells if v is not None)
        has_dates = any(isinstance(v, (datetime, dt_time)) for v in cells if v is not None)

        if non_empty > total_cols * 0.3 and (has_nums or has_dates):
            data_start = row_idx - 1
            break

    wb.close()
    return best_header, data_start


def _clean_column_name(name: str) -> str:
    r"""تنظيف اسم العمود — استخراج السطر العربي الأول فقط."""
    if not isinstance(name, str):
        return str(name)
    first_line = name.split("\n", maxsplit=1)[0].strip()
    return first_line or name.strip()


def _read_file() -> tuple[str, None] | tuple[None, tuple[pd.DataFrame, str]] | None:
    """طلب ملف إكسل وقراءته."""
    _show_step_header(1, "حدد ملف الإكسل")
    path = ask_file_dialog(
        title="اختر ملف الإكسل",
        filetypes=[("Excel", "*.xlsx *.xls"), ("All", "*.*")],
    )
    if not path:
        path = ask_file("مسار الملف:", extensions=[".xlsx", ".xls"])
    if not path:
        print_err("لم يتم اختيار ملف")
        return (NAV_BACK, None)

    try:
        xl = pd.ExcelFile(path, engine="openpyxl")
        sheets = xl.sheet_names
    except Exception as e:
        print_err(f"خطأ في فتح الملف: {e}")
        return (NAV_RETRY, None)

    if len(sheets) == 1:
        sheet = sheets[0]
        print_ok(f"الورقة: {sheet}")
    else:
        print_title("الأوراق المتوفرة:")
        for i, s in enumerate(sheets, 1):
            print(f"  {DIM}{i}{RESET}  {s}")
        choice = ask("رقم الورقة:")
        nav = _check_nav(choice)
        if nav:
            return (nav, None)
        try:
            idx = int(choice) - 1
            sheet = sheets[idx]
        except (ValueError, IndexError):
            print_err("اختيار غير صحيح")
            return (NAV_RETRY, None)

    try:
        header_row, data_row = _detect_header_row(path, sheet)
        skip_after = list(range(1, data_row - header_row))

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            df = pd.read_excel(
                path,
                sheet_name=sheet,
                header=header_row,
                skiprows=skip_after,
                engine="openpyxl",
            )

        df.columns = [_clean_column_name(c) for c in df.columns]
        unnamed = [c for c in df.columns if c.startswith("Unnamed")]
        if unnamed and len(unnamed) < len(df.columns):
            df = df.drop(columns=unnamed)
        df = df.dropna(how="all")

        if header_row > 0:
            print_ok(f"تم كشف العناوين تلقائياً في الصف {header_row + 1}")

    except Exception as e:
        print_err(f"خطأ في قراءة الورقة: {e}")
        return (NAV_RETRY, None)

    if df.empty:
        print_err("الورقة فارغة")
        return (NAV_RETRY, None)

    print_ok(f"تم قراءة {len(df)} سجل و {len(df.columns)} عمود")
    return (None, (df, os.path.basename(path), path))


def _detect_col_type(series: pd.Series) -> str:
    """كشف نوع العمود الفعلي."""
    from report_charts import detect_col_type
    return detect_col_type(series)


def _show_columns(df: pd.DataFrame) -> tuple[str, None] | tuple[None, list[str]]:
    """عرض الأعمدة وطلب اختيار."""
    _show_step_header(2, "الأعمدة المتوفرة")
    cols = list(df.columns)
    for i, col in enumerate(cols, 1):
        nuniq = df[col].nunique()
        col_type = _detect_col_type(df[col])
        print(f"  {DIM}{i}{RESET}  {col}")
        print(f"      {DIM}{nuniq} قيمة — {col_type}{RESET}")

    print_warn("اختر الأعمدة بالأرقام مفصولة بفواصل أو مسافات")
    print_warn("مثال: 1,3,5 أو 1 3 5 أو الكل لاختيار الكل")
    raw = ask("الأعمدة:")

    nav = _check_nav(raw)
    if nav:
        return (nav, None)

    if raw.strip() in ("الكل", "all"):
        selected = [c for c in cols if 1 < df[c].nunique() <= 50]
        if not selected:
            selected = cols[:5]
        print_ok(f"تم اختيار {len(selected)} عمود تلقائياً")
        return (None, selected)

    parts = raw.replace(",", " ").replace("،", " ").split()
    selected = []
    for p in parts:
        try:
            idx = int(p) - 1
            if 0 <= idx < len(cols):
                selected.append(cols[idx])
        except ValueError:
            continue

    if not selected:
        print_err("لم يتم اختيار أعمدة صحيحة")
        return (NAV_RETRY, None)

    print_ok(f"تم اختيار: {len(selected)} عمود")
    return (None, selected)


def _filter_columns(
    df: pd.DataFrame,
    columns: list[str],
) -> tuple[str, None] | tuple[None, tuple[pd.DataFrame, list[str]]]:
    """تصفية بيانات كل عمود."""
    _show_step_header(3, "تصفية بيانات الأعمدة")
    print_warn("لكل عمود: اختر الأرقام مفصولة بفواصل أو مسافات")
    print_warn("Enter أو 0 أو شامل = كل القيم")
    print()

    filtered_df = df.copy()
    kept_columns = []

    for col in columns:
        vc = filtered_df[col].dropna().value_counts()
        unique_vals = vc.index.tolist()

        if len(unique_vals) == 0:
            print(f"  {CYAN}{col}{RESET}")
            print(f"    {DIM}لا توجد بيانات — تم تخطيه{RESET}")
            continue

        if len(unique_vals) == 1:
            print(f"  {CYAN}{col}{RESET}")
            print(f"    {DIM}قيمة واحدة فقط: {unique_vals[0]} ({vc.iloc[0]}){RESET}")
            kept_columns.append(col)
            continue

        print(f"  {CYAN}{col}{RESET}")
        print(f"    {GREEN}0{RESET}  شامل ({len(unique_vals)} قيمة)")
        for idx, val in enumerate(unique_vals, 1):
            count = vc.iloc[idx - 1]
            print(f"    {DIM}{idx}{RESET}  {val} ({count})")
            if idx >= 20:
                remaining = len(unique_vals) - 20
                if remaining > 0:
                    print(f"    {DIM}... و {remaining} قيمة أخرى{RESET}")
                break

        raw = ask("اختيارك:")

        nav = _check_nav(raw)
        if nav:
            return (nav, None)

        cleaned = raw.strip()

        if not cleaned or cleaned == "0" or cleaned in ("شامل", "all"):
            print_ok("  شامل")
            kept_columns.append(col)
            continue

        parts = cleaned.replace(",", " ").replace("،", " ").split()
        chosen_vals = []
        for p in parts:
            try:
                idx = int(p) - 1
                if 0 <= idx < len(unique_vals):
                    chosen_vals.append(unique_vals[idx])
            except ValueError:
                continue

        if not chosen_vals:
            print_warn("  لم يُتعرف على الأرقام — شامل")
            kept_columns.append(col)
            continue

        filtered_df = filtered_df[filtered_df[col].isin(chosen_vals)]
        if len(chosen_vals) == 1:
            print_ok(f"  تم التصفية: {chosen_vals[0]} ({len(filtered_df)} سجل)")
        else:
            names = " + ".join(str(v) for v in chosen_vals)
            print_ok(f"  تم التصفية: {names} ({len(filtered_df)} سجل)")
        kept_columns.append(col)

        if filtered_df.empty:
            print_err("لا توجد سجلات بعد التصفية!")
            return (NAV_RETRY, None)

        print()

    if filtered_df.empty:
        print_err("لا توجد سجلات بعد التصفية!")
        return (NAV_RETRY, None)

    if not kept_columns:
        print_err("لم يبقَ أي عمود!")
        return (NAV_RETRY, None)

    print()
    print_ok(f"النتيجة: {len(filtered_df)} سجل و {len(kept_columns)} عمود")
    return (None, (filtered_df, kept_columns))


# أنواع الرسوم المتاحة
CHART_TYPES = {
    "1": ("bar_v", "أعمدة عمودية", "أشرطة رأسية لمقارنة القيم — الأنسب لعدد محدود من الفئات"),
    "2": ("bar_h", "أعمدة أفقية", "أشرطة أفقية — الأنسب عندما تكون أسماء الفئات طويلة"),
    "3": ("bar_stack", "أعمدة مكدسة", "أشرطة متراكمة فوق بعضها — تُظهر المجموع والتكوين معاً"),
    "4": ("pie", "دائري", "قطاعات دائرية — تُظهر حصة كل فئة من الإجمالي"),
    "5": ("donut", "دائري مجوّف", "مثل الدائري لكن مع فراغ في المنتصف لعرض الإجمالي"),
    "6": ("percent", "نسبة مئوية", "دائري يعرض النسب المئوية — للتركيز على الحصص"),
    "7": ("treemap", "خريطة شجرية", "مربعات متداخلة بحجم يتناسب مع القيمة — للبيانات الكثيرة"),
    "8": ("funnel", "قمع", "شكل قمعي تنازلي — يُظهر الترتيب من الأكبر للأصغر"),
    "9": ("waterfall", "شلال", "يُظهر كيف يتغير المجموع بإضافة أو خصم كل قيمة"),
    "10": ("gauge", "مؤشر أداء", "عداد نصف دائري — يُظهر القيمة الأعلى كنسبة من الإجمالي"),
    "11": ("kpi_cards", "بطاقات ملخص", "بطاقات رقمية تعرض: الإجمالي + عدد الفئات + الأعلى + الأدنى"),
    "12": ("sunburst", "شمسي", "حلقات متداخلة — مثل الدائري لكن بمستويات متعددة"),
}


def _ask_chart_types(
    columns: list[str],
) -> tuple[str, None] | tuple[None, list[tuple[str, str]]]:
    """لكل عمود مختار — يسأل عن نوع الرسم."""
    _show_step_header(4, "اختر نوع الرسم لكل عمود")
    print_title("أنواع الرسوم المتاحة:")
    for num, (_, name, desc) in CHART_TYPES.items():
        print(f"  {DIM}{num}{RESET}  {name}")
        print(f"      {DIM}{desc}{RESET}")
    print()
    print_warn("Enter فارغ = تخطي بدون رسم")
    print()

    selections = []
    for col in columns:
        print(f"  {CYAN}{col}{RESET}")
        raw = ask("نوع الرسم:")
        nav = _check_nav(raw)
        if nav:
            return (nav, None)

        if not raw.strip():
            print(f"    {DIM}تم التخطي{RESET}")
            continue

        parts = raw.replace(",", " ").replace("،", " ").split()
        for p in parts:
            p = p.strip()
            if p in CHART_TYPES:
                chart_type = CHART_TYPES[p][0]
                selections.append((col, chart_type))

    if not selections:
        print_warn("لم يتم اختيار أي رسم — سيتم إنشاء التقرير بالملخصات فقط")

    chart_count = len(selections)
    if chart_count > 0:
        print_ok(f"تم تحديد {chart_count} رسم بياني")
    return (None, selections)


def _execute_generation(ctx: dict) -> None:
    """تنفيذ توليد الرسوم وبناء PDF — مرحلة بدون تفاعل."""
    df = ctx["df"]
    df = normalize_code_columns(df)
    file_name = ctx["file_name"]
    selections = ctx["selections"] or []

    chart_dir = tempfile.mkdtemp(prefix="report_charts_")
    chart_paths = []
    if selections:
        print_header("جارٍ توليد الرسوم...")
        for i, (col, chart_type) in enumerate(selections):
            print_warn(f"  رسم {i + 1}/{len(selections)}: {col}")
            path = generate_chart(df, col, chart_type, chart_dir, i)
            if path:
                chart_paths.append(path)
        if chart_paths:
            print_ok(f"تم توليد {len(chart_paths)} رسم")

    chosen_columns = ctx.get("columns") or []
    df, chosen_columns = inject_dispatch_columns(df, chosen_columns)

    print_header("جارٍ توليد الملخصات...")
    summaries = auto_summaries(df, chosen_columns)
    if summaries:
        print_ok(f"تم توليد {len(summaries)} ملخص تلقائي")
    else:
        print_warn("لا توجد أعمدة نصية لإنشاء ملخصات")

    header_info = {}
    file_path = ctx.get("file_path")
    if file_path:
        header_info = _read_record_info(file_path)

    print_header("جارٍ بناء PDF...")
    pdf_path = build_pdf(
        df,
        chart_paths,
        selections,
        file_name,
        summaries,
        chosen_columns,
        header_info,
    )

    for p in chart_paths:
        if os.path.exists(p):
            os.remove(p)
    with contextlib.suppress(OSError):
        os.rmdir(chart_dir)

    if pdf_path and os.path.exists(pdf_path):
        print_ok("تم إنشاء التقرير بنجاح!")
        print(f"  {pdf_path}")
        if sys.platform == "win32":
            os.startfile(pdf_path)
    else:
        print_err("فشل إنشاء التقرير")


def _read_from_db():
    """قراءة بيانات الرحلات من القاعدة — بديل لقراءة إكسل."""
    project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if project_dir not in sys.path:
        sys.path.insert(0, project_dir)
    from db.queries import Q

    date = ask("التاريخ (مثل 2026-01-01):")
    if not date:
        return (NAV_BACK, None)

    with Q() as q:
        trips = q.trips_by_date(date)
        if not trips:
            print_err("لا توجد بيانات لهذا التاريخ!")
            return (NAV_RETRY, None)

    col_map = {
        "flight_number": "رقم الرحلة",
        "departure_time": "وقت الاقلاع",
        "passenger_count": "عدد الركاب",
        "destination": "الوجهة",
        "visa_type": "الفيزا",
        "campaign_name": "اسم الحملة",
        "status": "الحالة",
        "dispatch": "التفويج",
        "inspection": "الكشف",
        "shift_code": "النوبة",
        "source": "المصدر",
    }
    data = {ar: [t[en] for t in trips] for en, ar in col_map.items()}
    df = pd.DataFrame(data)

    filename = f"قاعدة_{date}"
    print_ok(f"تم قراءة {len(df)} سجل و {len(df.columns)} عمود من القاعدة")
    return (None, (df, filename, None))


def run() -> str:
    """نقطة الدخول — حلقة خطوات مع تنقل مرن."""
    print_header("منشئ التقارير")

    project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    db_path = os.path.join(project_dir, "db", "ht_sc.db")
    use_db = False
    if os.path.exists(db_path):
        print(f"  {CYAN}1{RESET} من ملف إكسل")
        print(f"  {CYAN}2{RESET} من قاعدة البيانات (أسرع)")
        choice = ask("اختر المصدر:")
        if choice == "2":
            use_db = True

    step = 1
    ctx = {
        "df": None,
        "df_filtered": None,
        "file_name": None,
        "file_path": None,
        "columns": None,
        "selections": None,
    }

    while True:
        if step == 1:
            if use_db:
                nav, data = _read_from_db()
            else:
                nav, data = _read_file()
            if nav == NAV_HOME:
                return "back"
            if nav == NAV_RETRY:
                continue
            if nav == NAV_BACK:
                return "back"
            ctx["df"], ctx["file_name"], ctx["file_path"] = data
            step = 2

        elif step == 2:
            nav, data = _show_columns(ctx["df"])
            if nav == NAV_HOME:
                return "back"
            if nav == NAV_RETRY:
                continue
            if nav == NAV_BACK:
                step = 1
                continue
            ctx["columns"] = data
            step = 3

        elif step == 3:
            nav, data = _filter_columns(ctx["df"], ctx["columns"])
            if nav == NAV_HOME:
                return "back"
            if nav == NAV_RETRY:
                continue
            if nav == NAV_BACK:
                step = 2
                continue
            ctx["df_filtered"], ctx["columns"] = data
            step = 4

        elif step == 4:
            nav, data = _ask_chart_types(ctx["columns"])
            if nav == NAV_HOME:
                return "back"
            if nav == NAV_RETRY:
                continue
            if nav == NAV_BACK:
                step = 3
                continue
            ctx["selections"] = data
            break

    if ctx["df_filtered"] is not None:
        ctx["df"] = ctx["df_filtered"]
    _execute_generation(ctx)
    pause()
    return "back"


if __name__ == "__main__":
    run()
