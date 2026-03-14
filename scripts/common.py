"""الدوال المشتركة بين جميع السكربتات."""

__all__ = [
    # ألوان وثوابت عرض
    "BLUE",
    "BOLD",
    "CYAN",
    "DIM",
    # ثوابت بيانات
    "FIELD_RES",
    "FIXED_COLS",
    "GREEN",
    "HEADER_RE",
    "IS_WT",
    "MAGENTA",
    "ORANGE",
    "PHOTO_RE",
    "RED",
    "RESET",
    "STAMP_RE",
    "STATUS_FAIL",
    "STATUS_OK",
    "STATUS_RUNNING",
    "TRIP_COLS",
    # دوال تفاعل
    "ask",
    "ask_continue",
    "ask_file",
    "ask_file_dialog",
    "ask_nav",
    # دوال عرض
    "bidi",
    # دوال بيانات
    "clean",
    "clear_screen",
    "date_to_tuple",
    "extract_photos",
    "extract_trips",
    "get_shift_name",
    "parse_messages",
    "parse_stamp",
    "pause",
    "print_breadcrumb",
    "print_err",
    "print_group",
    "print_header",
    "print_ok",
    "print_status",
    "print_status_header",
    "print_title",
    "print_warn",
    "ralign",
    # دوال إكسل
    "read_excel_data",
    "stamp_to_num",
    "strip_extra",
    "time_to_sec",
    # دوال مزامنة الإكسل مع القاعدة
    "sync_parse_folder_info",
    "sync_parse_file_info",
    "sync_get_or_create_sender",
    "sync_get_or_create_folder",
    "sync_get_or_create_file",
    "sync_excel_file",
    "sync_full",
    "sync_start_watcher",
    "sync_stop_watcher",
]

import builtins
import os
import re
import shutil


def _safe_log(_log_type, *args, **kwargs) -> None:
    """استدعاء دالة تسجيل بأمان - استيراد متأخر لتجنب الدوران."""
    try:
        from logger_config import log_action, log_file_op

        if _log_type == "action":
            log_action(*args, **kwargs)
        elif _log_type == "file_op":
            log_file_op(*args, **kwargs)
    except Exception:
        # التسجيل فشل — لا نوقف البرنامج
        pass


# === كشف نوع الطرفية ===
# Windows Terminal تعكس النص العربي تلقائياً
# لذلك نعكسه مسبقاً بـ get_display حتى يظهر سليماً بعد العكس المزدوج
IS_WT = "WT_SESSION" in os.environ
if IS_WT:
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display

        # الكائن الجاهز يُستخدم مباشرة بدل إنشاء واحد جديد كل استدعاء
        _reshaper = arabic_reshaper.default_reshaper
    except ImportError:
        IS_WT = False

# === حفظ الدوال الأصلية قبل التغليف ===
_orig_print = builtins.print
_orig_input = builtins.input

# === أنماط مُجمّعة ===
_ANSI_RE = re.compile(r"(\033\[[0-9;]*m)")
# نمط موحّد لكشف العربية - يغطي الكتلة الأساسية + أشكال العرض
_ARABIC_RE = re.compile(r"[\u0600-\u06FF\uFB50-\uFDFF\uFE70-\uFEFF]")
# أنماط تنظيف النصوص - مُجمّعة مرة واحدة بدل تجميعها كل استدعاء
_BIDI_RE = re.compile(r"[\u200f\u200e\u200b\u200c\u200d\u202a-\u202e\u2066-\u2069\ufeff]")
_STAMP_CLEAN_RE = re.compile(r"[\[\]\u200f\u200e\u200b\u200c\u200d\u202a-\u202e\u2066-\u2069\ufeff]")
_ATTACH_RE = re.compile(r"<المُرفق:.*?>")
_NO_IMAGE_RE = re.compile(r"لم يتم إدراج الصورة")

# === تخبئة عرض النافذة ===
# بدل سؤال النظام كل سطر, نحفظ القيمة نصف ثانية
import contextlib
import time as _time

_cached_cols = 0
_cached_cols_time = 0


def _vlen(text):
    """حساب العرض المرئي بدون أكواد ANSI"""
    clean = _ANSI_RE.sub("", text)
    return len(clean)


def bidi(text):
    """معالجة اتجاه النص للعرض الصحيح على الشاشة.

    الطرفية تعكس العربي تلقائياً - فنعكسه مسبقاً ليظهر سليماً
    """
    if not IS_WT:
        return text
    parts = _ANSI_RE.split(text)
    result = []
    for part in parts:
        if _ANSI_RE.match(part):
            # كود لون - يمر كما هو
            result.append(part)
        elif part:
            # نص عربي - يُعكس مسبقاً ليظهر سليماً بعد عكس الطرفية
            if _ARABIC_RE.search(part):
                # تشكيل الحروف أولاً (ربط متصل) ثم عكس الاتجاه
                result.append(get_display(_reshaper.reshape(part)))
            else:
                result.append(part)
    return "".join(result)


def ralign(text):
    """محاذاة النص إلى اليمين بإضافة مسافات.

    يعالج كل سطر على حدة عند وجود أسطر جديدة في النص
    """
    if not IS_WT or not text.strip():
        return text
    # تخبئة عرض النافذة - نسأل النظام مرة كل نصف ثانية فقط
    global _cached_cols, _cached_cols_time
    now = _time.time()
    if now - _cached_cols_time > 0.5:
        _cached_cols = shutil.get_terminal_size().columns
        _cached_cols_time = now
    width = _cached_cols
    # فصل الأسطر ومحاذاة كل سطر مستقل
    lines = text.split("\n")
    result = []
    for line in lines:
        vlen = _vlen(line)
        pad = width - vlen
        if pad > 0 and line.strip():
            result.append(" " * pad + line)
        else:
            # سطر فارغ أو أطول من الشاشة - يمر كما هو
            result.append(line)
    return "\n".join(result)


# === تغليف print و input لمعالجة العربية تلقائياً ===
# بهذا كل السكربتات تعمل بدون تعديل - المعالجة تحدث عند الطباعة
def _auto_print(*args, **kwargs) -> None:
    """تطبيق معالجة الاتجاه والمحاذاة تلقائياً على كل نص يُطبع"""
    new_args = []
    for arg in args:
        if isinstance(arg, str):
            processed = bidi(arg)
            # محاذاة تلقائية لليمين إذا يحتوي نص عربي ولم يُحاذى مسبقاً
            if _ARABIC_RE.search(arg) and kwargs.get("end", "\n") == "\n":
                processed = ralign(processed)
            new_args.append(processed)
        else:
            new_args.append(arg)
    _orig_print(*new_args, **kwargs)


def _auto_input(prompt=""):
    """تطبيق معالجة الاتجاه والمحاذاة على نص الطلب"""
    if isinstance(prompt, str):
        processed = bidi(prompt)
        # محاذاة نص الطلب لليمين إذا يحتوي عربي
        if _ARABIC_RE.search(prompt):
            processed = ralign(processed)
        return _orig_input(processed)
    return _orig_input(prompt)


# تفعيل التغليف في Windows Terminal فقط
if IS_WT:
    builtins.print = _auto_print
    builtins.input = _auto_input

# === ألوان الطرفية ===
GREEN = "\033[92m"
RED = "\033[91m"
ORANGE = "\033[93m"
CYAN = "\033[96m"
BOLD = "\033[1m"
DIM = "\033[2m"
RESET = "\033[0m"


def print_ok(msg) -> None:
    """طباعة رسالة نجاح بالأخضر"""
    # المحاذاة تتم تلقائياً عبر _auto_print
    print(f"{GREEN}{msg}{RESET}")


def print_err(msg) -> None:
    """طباعة رسالة خطأ بالأحمر"""
    print(f"{RED}{msg}{RESET}")


def print_warn(msg) -> None:
    """طباعة رسالة معالجة بالبرتقالي"""
    print(f"{ORANGE}{msg}{RESET}")


def print_title(msg) -> None:
    """طباعة عنوان ملون"""
    print(f"  {CYAN}{BOLD}{msg}{RESET}")


def print_header(msg) -> None:
    """طباعة رأس قسم ملون"""
    print(f"\n  {CYAN}{msg}{RESET}\n")


def ask(prompt):
    """طلب مدخل من المستخدم"""
    result = input(f"{GREEN}{prompt}{RESET} ").strip().strip('"')
    _safe_log(
        "action",
        f"إدخال مستخدم: {prompt}",
        action_type="input",
        prompt_text=prompt,
        user_response=result,
        script_name="common",
        func_name="ask",
    )
    return result


def ask_nav(prompt):
    """طلب مدخل مع إمكانية الإلغاء بـ ق أو ك.

    المخرجات: (أمر_تنقل, قيمة_المستخدم)
    - إذا أُلغي: ("cancel", None)
    - إذا أُدخلت قيمة: (None, "القيمة")
    """
    val = ask(prompt)
    # كشف أمر الإلغاء - ق بالعربية أو q بالإنجليزية
    if val.lower() in ("ق", "q"):
        return ("cancel", None)
    return (None, val)


def print_breadcrumb(current_name) -> None:
    """طباعة مسار التنقل في أعلى القائمة الفرعية"""
    print(f"  {DIM}القائمة الرئيسية ← {current_name}{RESET}")


def ask_file(prompt, must_exist=True, extensions=None):
    """طلب مسار ملف مع التحقق الكامل"""
    import os

    path = ask(prompt)
    if not path:
        print_err("لم يُدخل مسار!")
        _safe_log(
            "file_op",
            "لم يُدخل مسار",
            operation="open",
            file_path="",
            result="fail",
            error_reason="مسار فارغ",
            script_name="common",
            func_name="ask_file",
        )
        return None
    # إذا كان مجلداً: ابحث عن الملف المناسب بداخله
    if os.path.isdir(path):
        _safe_log(
            "file_op",
            f"المسار مجلد: {path}",
            operation="listdir",
            file_path=path,
            result="success",
            script_name="common",
            func_name="ask_file",
        )
        # بحث عن ملف واحد بالامتداد المطلوب
        found = None
        if extensions:
            for f in os.listdir(path):
                for ext in extensions:
                    if f.lower().endswith(ext):
                        if found:
                            # أكثر من ملف - لا يمكن تخمين المطلوب
                            print_err("المسار مجلد وليس ملف! يحتوي أكثر من ملف.")
                            print_err(f"حدد الملف مثلاً: {os.path.join(path, f)}")
                            return None
                        found = os.path.join(path, f)
        if found:
            print_warn(f"تم اكتشاف الملف: {os.path.basename(found)}")
            _safe_log(
                "file_op",
                f"اكتشاف ملف داخل مجلد: {found}",
                operation="exists",
                file_path=found,
                result="success",
                script_name="common",
                func_name="ask_file",
            )
            path = found
        else:
            # بحث عام: إذا يوجد ملف واحد فقط
            files = [f for f in os.listdir(path) if os.path.isfile(os.path.join(path, f))]
            if len(files) == 1:
                path = os.path.join(path, files[0])
                print_warn(f"تم اكتشاف الملف: {files[0]}")
            else:
                print_err("المسار مجلد وليس ملف!")
                if files:
                    print_err("الملفات الموجودة:")
                    for f in files[:5]:
                        print(f"  {f}")
                return None
    if must_exist and not os.path.exists(path):
        print_err(f"الملف غير موجود: {path}")
        _safe_log(
            "file_op",
            f"ملف غير موجود: {path}",
            operation="exists",
            file_path=path,
            result="fail",
            error_reason="غير موجود",
            script_name="common",
            func_name="ask_file",
        )
        return None
    # نجاح — تسجيل الملف المُختار
    _safe_log(
        "file_op",
        f"ملف مُختار: {path}",
        operation="open",
        file_path=path,
        result="success",
        script_name="common",
        func_name="ask_file",
    )
    return path


def pause() -> None:
    """انتظار الضغط على مفتاح"""
    input(f"\n{DIM}اضغط Enter للرجوع{RESET} ")
    _safe_log(
        "action",
        "ضغط Enter للرجوع",
        action_type="navigation",
        script_name="common",
        func_name="pause",
    )


def ask_continue(prompt="اكتب أمر أو Enter للرجوع"):
    """سؤال المستخدم - استمرار أو رجوع"""
    result = input(f"\n{DIM}{prompt}{RESET} ").strip()
    _safe_log(
        "action",
        f"قرار المستخدم: {result if result else 'رجوع'}",
        action_type="navigation",
        prompt_text=prompt,
        user_response=result or "(Enter)",
        script_name="common",
        func_name="ask_continue",
    )
    return result


# === نمط استخراج التاريخ والوقت من الطابع الزمني ===
STAMP_RE = re.compile(r"(\d+/\d+/\d+)[,,\s]+(\d+:\d+:\d+)")


def parse_stamp(raw):
    """استخراج التاريخ والوقت من طابع واتساب مثل [1/21/26, 22:52:40]."""
    cleaned = _STAMP_CLEAN_RE.sub("", raw).strip()
    m = STAMP_RE.search(cleaned)
    if not m:
        return None, None
    return m.group(1), m.group(2)


def clean(text):
    """تنظيف النص من علامات الاتجاه غير المرئية"""
    return _BIDI_RE.sub("", text).strip()


def strip_extra(text):
    """حذف علامات المرفقات ونص لم يتم إدراج الصورة"""
    text = _ATTACH_RE.sub("", text)
    text = _NO_IMAGE_RE.sub("", text)
    return text.strip()


def time_to_sec(t):
    """تحويل الوقت إلى ثوانٍ"""
    parts = t.split(":")
    return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])


def date_to_tuple(d):
    """تحويل تاريخ m/d/yy إلى ثلاثية للمقارنة"""
    parts = d.split("/")
    return (int(parts[2]), int(parts[0]), int(parts[1]))


def stamp_to_num(date_str, time_str):
    """تحويل تاريخ+وقت إلى رقم واحد قابل للمقارنة"""
    yy, mm, dd = date_to_tuple(date_str)
    return yy * 10_000_000_000 + mm * 100_000_000 + dd * 1_000_000 + time_to_sec(time_str)


# === نمط رأس رسالة الواتساب ===
HEADER_RE = re.compile(r"\[(\d+/\d+/\d+)[,,]\s*(\d+:\d+:\d+)\]\s*(.+?):\s*(.*)")

# === نمط استخراج اسم الصورة ===
PHOTO_RE = re.compile(r"<المُرفق:\s*([^>]+)>")

# === أنماط استخراج الحقول ===
FIELD_RES = {
    "رقم الرحلة": re.compile(r"رقم\s*الرحلة\s*:\s*(.*)"),
    "وقت الاقلاع": re.compile(r"وقت\s*الاقلاع\s*:\s*(.*)"),
    "عدد الركاب": re.compile(r"عدد\s*الركاب\s*:\s*(.*)"),
    "الوجهة": re.compile(r"الوجهة\s*:\s*(.*)"),
    "الفيزا": re.compile(r"الفيزا\s*:\s*(.*)"),
    "اسم الحملة": re.compile(r"(?:اسم\s+)?الحملة\s*:\s*(.*)"),
    "الحالة": re.compile(r"الحالة\s*:\s*(.*)"),
    "التفويج": re.compile(r"التفويج\s*:\s*(.*)"),
    "الكشف": re.compile(r"الكشف\s*:\s*(.*)"),
}

# === الأعمدة ===
FIXED_COLS = ["التاريخ", "الوقت", "المرسل", "الصور"]
TRIP_COLS = [
    "رقم الرحلة",
    "وقت الاقلاع",
    "عدد الركاب",
    "الوجهة",
    "الفيزا",
    "اسم الحملة",
    "الحالة",
    "التفويج",
    "الكشف",
]


def extract_photos(lines):
    """استخراج أسماء الصور من أسطر الرسالة"""
    photos = []
    for line in lines:
        for m in PHOTO_RE.finditer(line):
            name = m.group(1).strip()
            if name:
                photos.append(name)
    return photos


def parse_messages(filepath, start_date, start_time, end_date, end_time):
    """قراءة الملف واستخراج الرسائل في النطاق الزمني"""
    with open(filepath, encoding="utf-8") as f:
        lines = f.readlines()

    start_num = stamp_to_num(start_date, start_time)
    end_num = stamp_to_num(end_date, end_time)

    messages = []
    current = None
    in_range = False

    for raw_line in lines:
        line = clean(raw_line)
        if not line:
            continue

        m = HEADER_RE.search(line)
        if m:
            if current and in_range:
                messages.append(current)

            date = m.group(1)
            time_str = m.group(2)
            msg_num = stamp_to_num(date, time_str)
            in_range = start_num <= msg_num <= end_num

            if in_range:
                content = m.group(4).strip()
                current = {
                    "date": date,
                    "time": time_str,
                    "sender": m.group(3).strip(),
                    "lines": [content] if content else [],
                    "raw_lines": [line],
                }
            else:
                current = None
        elif current and in_range:
            current["lines"].append(line)
            current["raw_lines"].append(line)

    if current and in_range:
        messages.append(current)

    return messages


def extract_trips(msg):
    """استخراج رحلة أو أكثر من رسالة واحدة"""
    if not msg["lines"]:
        return None

    full_text = " ".join(msg["lines"])
    if "تم حذف هذه الرسالة" in full_text:
        return None

    has_data = False
    for l in msg["lines"]:
        s = strip_extra(l)
        if s and "لم يتم" not in s and "<المُرفق" not in s:
            has_data = True
            break
    if not has_data:
        return None

    flight_re = FIELD_RES["رقم الرحلة"]
    blocks = []
    current_block = []

    for line in msg["lines"]:
        line_clean = strip_extra(line)
        if not line_clean:
            continue
        if flight_re.search(line_clean) and current_block:
            blocks.append(current_block)
            current_block = [line]
        else:
            current_block.append(line)

    if current_block:
        blocks.append(current_block)

    if not blocks:
        return None

    trips = []
    for block in blocks:
        trip = {}
        tafweej_list = []

        for line in block:
            line_clean = strip_extra(line)
            if not line_clean:
                continue

            for field_name, regex in FIELD_RES.items():
                match = regex.search(line_clean)
                if match:
                    value = strip_extra(match.group(1)).strip()
                    if field_name == "التفويج":
                        tafweej_list.append(value)
                    elif field_name not in trip:
                        trip[field_name] = value
                    break

        if tafweej_list:
            if "الكشف" in trip:
                trip["التفويج"] = tafweej_list[0]
            elif len(tafweej_list) >= 2:
                trip["التفويج"] = tafweej_list[0]
                trip["الكشف"] = tafweej_list[1]
            else:
                trip["التفويج"] = tafweej_list[0]

        if "رقم الرحلة" in trip:
            trips.append(trip)

    return trips or None


def read_excel_data(path, header_row=1):
    """قراءة ملف إكسل بوضع القراءة فقط - حماية الذاكرة.

    المدخلات: مسار الملف + رقم سطر الرؤوس (افتراضي 1)
    المخرجات: (headers, rows) أو (None, None) عند الفشل
    - headers = قائمة أسماء الأعمدة كنصوص
    - rows = قائمة صفوف, كل صف = tuple من القيم
    """
    import openpyxl

    # حجم الملف للتسجيل
    _fsize = None
    with contextlib.suppress(Exception):
        _fsize = os.path.getsize(path)

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print_err(f"فتح الملف: {e}")
        _safe_log(
            "file_op",
            f"فشل فتح إكسل: {e}",
            operation="read",
            file_path=path,
            file_size_bytes=_fsize,
            result="fail",
            error_reason=str(e),
            script_name="common",
            func_name="read_excel_data",
        )
        return None, None

    try:
        ws = wb.active
        # قراءة الرؤوس - iter_rows لأن ws[n] لا يعمل مع read_only
        header_cells = next(ws.iter_rows(min_row=header_row, max_row=header_row))
        headers = [str(c.value).strip() if c.value else "" for c in header_cells]
        # قراءة كل الصفوف كقيم مباشرة
        rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
    finally:
        wb.close()

    # تسجيل نجاح القراءة مع التفاصيل
    _safe_log(
        "file_op",
        f"قراءة إكسل: {len(headers)} عمود × {len(rows)} صف",
        operation="read",
        file_path=path,
        file_size_bytes=_fsize,
        result="success",
        script_name="common",
        func_name="read_excel_data",
    )

    return headers, rows


def ask_file_dialog(title="اختر الملف", filetypes=None):
    """فتح نافذة اختيار ملف - تسقط بهدوء إذا tkinter غير متوفرة"""
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError:
        # tkinter غير مثبتة - نستخدم الإدخال النصي بدلاً منها
        print_warn("نافذة الاختيار غير متوفرة - أدخل المسار يدوياً")
        _safe_log(
            "action",
            "فشل فتح نافذة الاختيار — tkinter غير متوفرة",
            action_type="cancel",
            script_name="common",
            func_name="ask_file_dialog",
        )
        return ask_file(title)

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    if filetypes is None:
        filetypes = [("جميع الملفات", "*.*")]

    filepath = filedialog.askopenfilename(title=title, filetypes=filetypes)
    root.destroy()

    if filepath:
        _safe_log(
            "action",
            f"اختيار ملف من النافذة: {filepath}",
            action_type="input",
            prompt_text=title,
            user_response=filepath,
            is_valid=1,
            script_name="common",
            func_name="ask_file_dialog",
        )
    else:
        _safe_log(
            "action",
            "إلغاء نافذة اختيار الملف",
            action_type="cancel",
            prompt_text=title,
            user_response="(إلغاء)",
            is_valid=0,
            script_name="common",
            func_name="ask_file_dialog",
        )

    return filepath or None


# === ألوان إضافية ===
BLUE = "\033[94m"
MAGENTA = "\033[95m"

# === ثوابت حالة التنفيذ ===
STATUS_RUNNING = f"{ORANGE}...{RESET}"
STATUS_OK = f"{GREEN}OK{RESET}"
STATUS_FAIL = f"{RED}FAIL{RESET}"


def clear_screen() -> None:
    """مسح الشاشة باستخدام تسلسل ANSI - يعمل في Git Bash على ويندوز"""
    # cls لا يعمل في Git Bash لذلك نستخدم ANSI escape
    print("\033[2J\033[H", end="")
    _safe_log(
        "action",
        "مسح الشاشة",
        action_type="navigation",
        script_name="common",
        func_name="clear_screen",
    )


def get_shift_name():
    """تحديد النوبة الحالية حسب الوقت - بداية اليوم 5:45 صباحاً"""
    from datetime import datetime

    now = datetime.now()
    hour = now.hour
    minute = now.minute
    total_min = hour * 60 + minute
    # A صباح 5:45-13:44 / B ظهر 13:45-21:44 / C مساء 21:45-5:44
    if 345 <= total_min < 825:
        return "A", "صباح"
    if 825 <= total_min < 1305:
        return "B", "ظهر"
    return "C", "مساء"


def print_status_header() -> None:
    """سطر حالة: وقت + نوبة + ذاكرة"""
    from datetime import datetime

    now = datetime.now()
    time_str = now.strftime("%H:%M")

    shift_code, shift_name = get_shift_name()

    # الذاكرة المتاحة
    try:
        import psutil

        mem = psutil.virtual_memory()
        avail_mb = mem.available // (1024 * 1024)
        mem_text = f"{avail_mb}"
        if avail_mb > 500:
            mem_color = GREEN
        elif avail_mb > 300:
            mem_color = ORANGE
        else:
            mem_color = RED
    except ImportError:
        mem_text = "?"
        mem_color = DIM

    # عرض بسيط: وقت | نوبة | ذاكرة
    line = f"{CYAN}{time_str}{RESET} | {GREEN}{shift_code} {shift_name}{RESET} | {mem_color}{mem_text} م.ب{RESET}"
    # المحاذاة تتم تلقائياً عبر _auto_print
    print(f"  {line}")
    print()


def print_group(title, color, items) -> None:
    """عرض مجموعة من الخيارات.

    title: اسم المجموعة
    color: لون العنوان
    items: قائمة من (رقم, وصف)
    """
    print(f"  {color}{BOLD}{title}{RESET}")
    for num, desc in items:
        print(f"    {DIM}{num}{RESET}  {desc}")
    print()


def print_status(status_text) -> None:
    """طباعة حالة التنفيذ الملونة"""
    print(f"\n  {status_text}")


# ══════════════════════════════════════════════════════
# دوال مزامنة ملفات الإكسل مع قاعدة البيانات
# ══════════════════════════════════════════════════════

import logging
import sqlite3
from pathlib import Path
from datetime import datetime

import openpyxl
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# إعداد سجل المزامنة
_sync_logger = logging.getLogger("excel_sync")
if not _sync_logger.handlers:
    _sync_handler = logging.StreamHandler()
    _sync_handler.setFormatter(
        logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%H:%M:%S")
    )
    _sync_logger.addHandler(_sync_handler)
    _sync_logger.setLevel(logging.INFO)

# مسار مجلد السجلات
_SYNC_RECORD_DIR = Path(__file__).parent.parent / "monitor" / "Record_2026"

# مسار القاعدة
_SYNC_DB_DIR = Path(__file__).parent.parent / "db"
_SYNC_DB_PATH = _SYNC_DB_DIR / "ht_sc.db"

# استيراد الثوابت المشتركة من db.constants
import sys
_PROJECT_ROOT = str(Path(__file__).parent.parent)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from db.constants import MAX_TRIPS as _SYNC_MAX_TRIPS
from db.constants import TRIP_FIELDS as _SYNC_TRIP_FIELDS
from db.helpers import build_records_insert_sql

_SYNC_INSERT_SQL = build_records_insert_sql()

# مرجع المراقب — يُستخدم لإيقافه لاحقاً
_sync_observer = None


def _sync_get_connection(db_path: str | None = None) -> sqlite3.Connection:
    """اتصال بالقاعدة مع تفعيل المفاتيح الأجنبية."""
    path = db_path or str(_SYNC_DB_PATH)
    conn = sqlite3.connect(path)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode=WAL")
    conn.row_factory = sqlite3.Row
    return conn


def sync_parse_folder_info(folder_path: Path) -> dict | None:
    """استخراج معلومات الشهر واليوم من مسار المجلد."""
    name = folder_path.name
    if not re.match(r"^\d{2}-\d{2}-\d{4}$", name):
        return None
    parent_name = folder_path.parent.name
    if not re.match(r"^\d{2}-\d{4}$", parent_name):
        return None
    return {"month": parent_name, "day": name, "path": str(folder_path)}


def sync_parse_file_info(file_path: Path) -> dict | None:
    """استخراج معلومات النوبة والنوع من اسم ملف الإكسل."""
    name = file_path.name
    if not name.endswith(".xlsx") or name.endswith(".tmp"):
        return None
    match = re.match(
        r"^\d{1,2}-\d{2}-\d{4}-Record-([ABC])(-Stuck)?\.xlsx$", name
    )
    if not match:
        return None
    return {
        "filename": name,
        "shift_code": match.group(1),
        "file_type": "stuck" if match.group(2) is not None else "normal",
    }


def sync_get_or_create_sender(
    conn: sqlite3.Connection, name: str
) -> int:
    """استرجاع معرّف المرسل أو إنشاؤه."""
    if not name or not re.search(r"[\u0600-\u06FFa-zA-Z]{2,}", name):
        name = "غير معروف"
    row = conn.execute(
        "SELECT id FROM senders WHERE name = ?", (name,)
    ).fetchone()
    if row:
        return row[0]
    cur = conn.execute(
        "INSERT INTO senders (name) VALUES (?)", (name,)
    )
    return cur.lastrowid


def sync_get_or_create_folder(
    conn: sqlite3.Connection, info: dict
) -> int:
    """استرجاع معرّف المجلد أو إنشاؤه."""
    row = conn.execute(
        "SELECT id FROM excel_folders WHERE day = ?", (info["day"],)
    ).fetchone()
    if row:
        return row[0]
    cur = conn.execute(
        "INSERT INTO excel_folders (month, day, path) VALUES (?, ?, ?)",
        (info["month"], info["day"], info["path"]),
    )
    return cur.lastrowid


def sync_get_or_create_file(
    conn: sqlite3.Connection, folder_id: int, info: dict
) -> int:
    """استرجاع معرّف الملف أو إنشاؤه."""
    row = conn.execute(
        "SELECT id FROM excel_files WHERE folder_id = ? AND filename = ?",
        (folder_id, info["filename"]),
    ).fetchone()
    if row:
        return row[0]
    cur = conn.execute(
        "INSERT INTO excel_files "
        "(folder_id, shift_code, filename, file_type) "
        "VALUES (?, ?, ?, ?)",
        (folder_id, info["shift_code"], info["filename"], info["file_type"]),
    )
    return cur.lastrowid


def sync_excel_file(
    conn: sqlite3.Connection, file_path: Path
) -> int:
    """مزامنة ملف إكسل واحد — إرجاع عدد الصفوف الجديدة."""
    folder_path = file_path.parent
    folder_info = sync_parse_folder_info(folder_path)
    if not folder_info:
        return 0
    file_info = sync_parse_file_info(file_path)
    if not file_info:
        return 0

    folder_id = sync_get_or_create_folder(conn, folder_info)
    file_id = sync_get_or_create_file(conn, folder_id, file_info)

    existing_count = conn.execute(
        "SELECT COUNT(*) FROM records WHERE file_id = ?", (file_id,)
    ).fetchone()[0]

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
    except Exception as e:
        _sync_logger.error(f"فشل قراءة الملف {file_path.name}: {e}")
        return 0

    ws = wb.active
    if ws is None or ws.max_row is None or ws.max_row < 2:
        wb.close()
        return 0

    total_rows = ws.max_row - 1
    new_rows = 0

    if total_rows <= existing_count:
        wb.close()
        return 0

    col_count = ws.max_column
    date_str = folder_info["day"]
    start_row = existing_count + 2

    # كشف التنسيق: الجديد = المرسل في العمود 2، القديم = في آخر الصف
    header_2 = str(ws.cell(1, 2).value or "")
    is_new_format = "المرسل" in header_2 or "اسم" in header_2

    for row_idx in range(start_row, ws.max_row + 1):
        cells = []
        for col in range(1, col_count + 1):
            cells.append(ws.cell(row_idx, col).value)

        if not any(c is not None for c in cells):
            continue

        # استخراج المرسل والوقت حسب التنسيق
        if is_new_format:
            # الجديد: [0] نوبة، [1] مرسل، [2] وقت، [3+] رحلات
            sender_name = str(cells[1]).strip() if len(cells) > 1 and cells[1] else ""
            msg_time = str(cells[2]).strip() if len(cells) > 2 and cells[2] else ""
            trip_start = 3
        else:
            # القديم: [0] نوبة، [1+] رحلات، [آخر 2] وقت + مرسل
            last_idx = 0
            for i in range(len(cells) - 1, -1, -1):
                if cells[i] is not None:
                    last_idx = i
                    break
            effective_len = last_idx + 1
            row_trips_old = max(1, (effective_len - 3) // 9)
            mt_idx = 1 + (row_trips_old * 9)
            sn_idx = mt_idx + 1
            msg_time = str(cells[mt_idx]).strip() if mt_idx < len(cells) and cells[mt_idx] else ""
            sender_name = str(cells[sn_idx]).strip() if sn_idx < len(cells) and cells[sn_idx] else ""
            trip_start = 1

        # حساب عدد الرحلات
        remaining = col_count - trip_start
        max_possible = remaining // 9
        trip_count = 0
        for t in range(min(max_possible, _SYNC_MAX_TRIPS)):
            base = trip_start + (t * 9)
            trip_slice = cells[base:base + 9]
            if any(c is not None for c in trip_slice):
                trip_count = t + 1
        if trip_count == 0:
            trip_count = 1

        # بناء قيم الرحلات (9 حقول × 10 رحلات = 90 قيمة)
        trip_values = []
        for t in range(_SYNC_MAX_TRIPS):
            base = trip_start + (t * 9)
            for i, field in enumerate(_SYNC_TRIP_FIELDS):
                idx = base + i
                if t < trip_count and idx < len(cells) and cells[idx] is not None:
                    if field == "passenger_count":
                        val = str(cells[idx]).strip()
                        nums = re.sub(r"[^\d]", "", val)
                        trip_values.append(int(nums) if nums else 0)
                    else:
                        trip_values.append(str(cells[idx]).strip())
                else:
                    trip_values.append(0 if field == "passenger_count" else None)

        fixed_values = [
            date_str,
            file_info["shift_code"],
            sender_name,
            msg_time,
            row_idx - 1,
            trip_count,
            file_id,
        ]

        conn.execute(_SYNC_INSERT_SQL, fixed_values + trip_values)
        new_rows += 1

    conn.execute(
        "UPDATE excel_files SET row_count = ?, last_synced = ? "
        "WHERE id = ?",
        (total_rows, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), file_id),
    )
    conn.commit()
    wb.close()

    if new_rows > 0:
        _sync_logger.info(
            f"مزامنة {file_path.name}: {new_rows} صف جديد"
        )
    return new_rows


def sync_full(conn: sqlite3.Connection) -> int:
    """مزامنة كاملة لجميع ملفات الإكسل الموجودة."""
    if not _SYNC_RECORD_DIR.exists():
        _sync_logger.warning(f"المجلد غير موجود: {_SYNC_RECORD_DIR}")
        return 0

    total = 0
    for xlsx_file in sorted(_SYNC_RECORD_DIR.rglob("*.xlsx")):
        if xlsx_file.name.endswith(".tmp"):
            continue
        total += sync_excel_file(conn, xlsx_file)

    _sync_logger.info(f"المزامنة الكاملة: {total} صف جديد")
    return total


class _SyncExcelHandler(FileSystemEventHandler):
    """مراقب أحداث الملفات — يزامن عند إنشاء أو تعديل ملف إكسل."""

    def __init__(self, db_path: str):
        """تهيئة المراقب مع مسار القاعدة."""
        super().__init__()
        self.db_path = db_path

    def _handle(self, event) -> None:
        """معالجة حدث ملف."""
        if event.is_directory:
            return
        file_path = Path(event.src_path)
        if not file_path.name.endswith(".xlsx"):
            return
        if file_path.name.endswith(".tmp"):
            return
        _time.sleep(0.5)
        try:
            conn = _sync_get_connection(self.db_path)
            conn.execute("PRAGMA busy_timeout = 5000")
            sync_excel_file(conn, file_path)
            conn.close()
        except Exception as e:
            _sync_logger.error(f"خطأ مزامنة {file_path.name}: {e}")

    def on_created(self, event) -> None:
        """عند إنشاء ملف جديد."""
        self._handle(event)

    def on_modified(self, event) -> None:
        """عند تعديل ملف موجود."""
        self._handle(event)

    def on_moved(self, event) -> None:
        """عند إعادة تسمية ملف."""
        if not hasattr(event, "dest_path"):
            return
        dest = Path(event.dest_path)
        if not dest.name.endswith(".xlsx"):
            return
        if dest.name.endswith(".tmp"):
            return
        try:
            conn = _sync_get_connection(self.db_path)
            conn.execute("PRAGMA busy_timeout = 5000")
            sync_excel_file(conn, dest)
            conn.close()
        except Exception as e:
            _sync_logger.error(f"خطأ مزامنة {dest.name}: {e}")


def sync_start_watcher(db_path: str | None = None) -> bool:
    """تشغيل مراقب المزامنة التلقائية — يُرجع نجاح التشغيل."""
    global _sync_observer
    if _sync_observer is not None:
        _sync_logger.warning("المزامنة تعمل مسبقاً")
        return True

    db = db_path or str(_SYNC_DB_PATH)

    if not _SYNC_RECORD_DIR.exists():
        _sync_logger.error(f"المجلد غير موجود: {_SYNC_RECORD_DIR}")
        return False

    # مزامنة كاملة أولاً
    try:
        conn = _sync_get_connection(db)
        conn.execute("PRAGMA busy_timeout = 5000")
        sync_full(conn)
        conn.close()
    except Exception as e:
        _sync_logger.error(f"فشل المزامنة الأولية: {e}")
        return False

    # تشغيل المراقب
    handler = _SyncExcelHandler(db)
    _sync_observer = Observer()
    _sync_observer.schedule(handler, str(_SYNC_RECORD_DIR), recursive=True)
    _sync_observer.daemon = True
    _sync_observer.start()
    _sync_logger.info(f"مراقب المزامنة يعمل على: {_SYNC_RECORD_DIR}")
    return True


def sync_stop_watcher() -> None:
    """إيقاف مراقب المزامنة."""
    global _sync_observer
    if _sync_observer is None:
        return
    _sync_observer.stop()
    _sync_observer.join(timeout=5)
    _sync_observer = None
    _sync_logger.info("تم إيقاف مراقب المزامنة")
