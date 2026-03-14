"""مولد التقارير والإحصائيات.

يقرأ بيانات من ملف RECORD ويملأ قوالب ODS عبر LibreOffice
التنسيق والألوان والتصميم محفوظة بالكامل
"""

import contextlib
import os
import shutil
import subprocess
import tempfile
import time as _time
from datetime import date, datetime

import openpyxl

from common import (
    BOLD,
    CYAN,
    DIM,
    GREEN,
    RED,
    RESET,
    ask,
    print_err,
    print_header,
    print_ok,
    print_warn,
)
from logger_config import (
    log_action,
    log_exception,
    log_file_op,
    log_processing,
    log_quality,
)

# ──────────────────────────────────────────
# الثوابت
# ──────────────────────────────────────────

# مسار القوالب
TEMPLATES_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "15_16_2026_\u0627\u062d\u0635\u0627\u0626\u064a\u0629",
    "Reort_Form",
)

# مسار LibreOffice
SOFFICE = "C:/Program Files/LibreOffice/program/soffice.exe"

# خرائط الخلايا: خلية_القالب ← خلية_المصدر
MAPPINGS = {
    "A": {
        "A2": "AR9",
        "C3": "N3",
        "C4": "O3",
        "C5": "Y3",
        "C6": "AA3",
        "C8": "P3",
        "C9": "Z3",
        "C10": "AB3",
        "C17": "Q3",
        "C18": "R3",
        "C19": "T3",
        "C20": "S3",
        "C21": "U3",
    },
    "B": {
        "A2": "AR9",
        "C3": "N4",
        "C4": "O4",
        "C5": "Y4",
        "C6": "AA4",
        "C8": "P4",
        "C9": "Z4",
        "C10": "AB4",
        "C17": "Q4",
        "C18": "R4",
        "C19": "T4",
        "C20": "S4",
        "C21": "U4",
    },
    "C": {
        "A2": "AR9",
        "C3": "N5",
        "C4": "O5",
        "C5": "Y5",
        "C6": "AA5",
        "C8": "P5",
        "C9": "Z5",
        "C10": "AB5",
        "C17": "Q5",
        "C18": "R5",
        "C19": "T5",
        "C20": "S5",
        "C21": "U5",
    },
    "OneDay": {
        "A2": "AR9",
        # نوبة A - العمود C (B=هيدر)
        "C3": "N3",
        "C4": "O3",
        "C5": "Y3",
        "C6": "AA3",
        "C8": "P3",
        "C9": "Z3",
        "C10": "AB3",
        "C16": "Q3",
        "C17": "R3",
        "C18": "T3",
        "C19": "S3",
        "C20": "U3",
        # نوبة B - العمود D
        "D3": "N4",
        "D4": "O4",
        "D5": "Y4",
        "D6": "AA4",
        "D8": "P4",
        "D9": "Z4",
        "D10": "AB4",
        "D16": "Q4",
        "D17": "R4",
        "D18": "T4",
        "D19": "S4",
        "D20": "U4",
        # نوبة C - العمود E
        "E3": "N5",
        "E4": "O5",
        "E5": "Y5",
        "E6": "AA5",
        "E8": "P5",
        "E9": "Z5",
        "E10": "AB5",
        "E16": "Q5",
        "E17": "R5",
        "E18": "T5",
        "E19": "S5",
        "E20": "U5",
    },
}

# خلايا المعادلات - ممنوع الكتابة فوقها
FORMULAS = {
    "A": {"C16"},
    "B": {"C16"},
    "C": {"C16"},
    "OneDay": {"F4", "F5", "F6", "F8", "F9", "F10", "F16", "F17", "F18", "F19", "F20"},
}

# معادلات تُضاف برمجياً - غير موجودة في القالب الأصلي
INJECT_FORMULAS = {
    "OneDay": {
        "F4": "=SUM(C4:E4)",
        "F5": "=SUM(C5:E5)",
        "F6": "=SUM(C6:E6)",
        "F8": "=SUM(C8:E8)",
        "F9": "=SUM(C9:E9)",
        "F10": "=SUM(C10:E10)",
        "F16": "=SUM(C16:E16)",
        "F17": "=SUM(C17:E17)",
        "F18": "=SUM(C18:E18)",
        "F19": "=SUM(C19:E19)",
        "F20": "=SUM(C20:E20)",
    },
}


def _format_size(size_bytes) -> str:
    """تحويل الحجم لنص مقروء"""
    if size_bytes < 1024:
        return f"{size_bytes} ب"
    if size_bytes < 1024 * 1024:
        return f"{size_bytes // 1024} ك.ب"
    return f"{size_bytes // (1024 * 1024)} م.ب"


# ──────────────────────────────────────────
# 1. تصفح المسارات واختيار الملف
# ──────────────────────────────────────────


def _browse_to_file(start_path):
    """تصفح تفاعلي حتى اختيار ملف إكسل - يُرجع المسار الكامل أو None"""
    path = start_path
    while True:
        try:
            entries = sorted(os.listdir(path))
        except PermissionError:
            print_err("لا توجد صلاحية!")
            log_file_op(
                "فشل الوصول للمجلد",
                operation="listdir",
                file_path=path,
                result="fail",
                error_reason="PermissionError",
                script_name="reports_generator",
                func_name="_browse_to_file",
            )
            return None

        # جمع المجلدات + ملفات إكسل فقط
        dirs = [e for e in entries if os.path.isdir(os.path.join(path, e))]
        xlsx = [
            e
            for e in entries
            if os.path.isfile(os.path.join(path, e)) and e.lower().endswith((".xlsx", ".xls")) and not e.startswith("~")
        ]

        # عرض المسار الحالي
        print(f"\n  {BOLD}{os.path.basename(path) or path}{RESET}/")

        items = [("dir", d) for d in dirs]
        items.extend(("file", f) for f in xlsx)

        if not items:
            print_warn("المجلد فارغ")
            # رجوع تلقائي
            parent = os.path.dirname(path)
            if parent == path:
                return None
            path = parent
            continue

        for i, (etype, name) in enumerate(items, 1):
            if etype == "dir":
                print(f"    {CYAN}{i}{RESET}  {name}/")
            else:
                full = os.path.join(path, name)
                try:
                    size = _format_size(os.path.getsize(full))
                except OSError:
                    size = "?"
                print(f"    {GREEN}{i}{RESET}  {name}  {DIM}{size}{RESET}")

        print(f"    {RED}0{RESET}  رجوع")

        choice = ask("اختر:")
        if not choice or choice == "0":
            # رجوع مستوى واحد
            parent = os.path.dirname(path)
            if parent == path or parent == os.path.dirname(start_path):
                # وصلنا لأعلى من نقطة البداية
                return None
            path = parent
            continue

        if choice.isdigit() and 1 <= int(choice) <= len(items):
            etype, name = items[int(choice) - 1]
            full = os.path.join(path, name)
            if etype == "dir":
                log_action(
                    f"دخول مجلد: {name}",
                    action_type="navigation",
                    user_response=choice,
                    script_name="reports_generator",
                    func_name="_browse_to_file",
                )
                path = full
            else:
                # اختار ملف إكسل
                log_action(
                    f"اختيار ملف: {name}",
                    action_type="input",
                    user_response=full,
                    is_valid=1,
                    script_name="reports_generator",
                    func_name="_browse_to_file",
                )
                return full
        else:
            print_err("اختيار غير صحيح!")


def _select_path(paths):
    """عرض المسارات المحفوظة واختيار واحد"""
    print_header("اختر مسار البيانات")
    for i, entry in enumerate(paths, 1):
        name = entry.get("name", "")
        epath = entry.get("path", "")
        etype = entry.get("type", "dir")
        if etype == "dir":
            print(f"    {CYAN}{i}{RESET}  {name}/")
        else:
            print(f"    {GREEN}{i}{RESET}  {name}")
        print(f"       {DIM}{epath}{RESET}")
    print(f"    {RED}0{RESET}  رجوع")

    choice = ask("اختر:")
    if not choice or choice == "0":
        return None
    if choice.isdigit() and 1 <= int(choice) <= len(paths):
        selected = paths[int(choice) - 1].get("path", "")
        log_action(
            f"اختيار مسار: {selected}",
            action_type="menu_choice",
            prompt_text="اختر:",
            user_response=choice,
            is_valid=1,
            script_name="reports_generator",
            func_name="_select_path",
        )
        return selected
    print_err("اختيار غير صحيح!")
    return None


# ──────────────────────────────────────────
# 2. اختيار نوع التقرير
# ──────────────────────────────────────────


def _select_report_type():
    """عرض 4 خيارات للتقارير"""
    print_header("اختر نوع التقرير")
    print(f"    {CYAN}1{RESET}  نوبة A")
    print(f"    {CYAN}2{RESET}  نوبة B")
    print(f"    {CYAN}3{RESET}  نوبة C")
    print(f"    {CYAN}4{RESET}  يومية (جميع النوبات)")
    print(f"    {RED}0{RESET}  رجوع")

    choice = ask("اختر:")
    types = {"1": "A", "2": "B", "3": "C", "4": "OneDay"}
    if not choice or choice == "0":
        return None
    result = types.get(choice)
    if not result:
        print_err("اختيار غير صحيح!")
    else:
        log_action(
            f"اختيار نوع التقرير: {result}",
            action_type="menu_choice",
            prompt_text="اختر:",
            user_response=choice,
            is_valid=1,
            script_name="reports_generator",
            func_name="_select_report_type",
        )
    return result


# ──────────────────────────────────────────
# 3. قراءة البيانات من الملف المصدر
# ──────────────────────────────────────────


def _read_source(source_file, cell_mapping):
    """قراءة القيم المطلوبة من ملف RECORD"""
    # حجم الملف للتسجيل
    _fsize = None
    with contextlib.suppress(Exception):
        _fsize = os.path.getsize(source_file)
    try:
        wb = openpyxl.load_workbook(source_file, data_only=True)
        ws = wb.active
        data = {}
        for template_cell, source_cell in cell_mapping.items():
            val = ws[source_cell].value
            if isinstance(val, datetime):
                val = val.date()
            data[template_cell] = val
        wb.close()
        log_file_op(
            f"قراءة RECORD: {len(data)} خلية",
            operation="read",
            file_path=source_file,
            file_size_bytes=_fsize,
            result="success",
            script_name="reports_generator",
            func_name="_read_source",
        )
        return data
    except FileNotFoundError:
        print_err("ملف البيانات غير موجود!")
        log_exception(
            "ملف RECORD غير موجود",
            script_name="reports_generator",
            func_name="_read_source",
            context=source_file,
        )
    except PermissionError:
        print_err("الملف مفتوح في برنامج آخر!")
        log_exception(
            "ملف RECORD مقفل",
            script_name="reports_generator",
            func_name="_read_source",
            context=source_file,
        )
    except Exception as e:
        print_err(f"خطأ في القراءة: {e}")
        log_exception(
            f"خطأ قراءة RECORD: {e}",
            exc=e,
            script_name="reports_generator",
            func_name="_read_source",
        )
    return None


# ──────────────────────────────────────────
# 4. نسخ القالب وتحويله بـ LibreOffice
# ──────────────────────────────────────────


def _convert_template(report_type, output_dir):  # noqa: ARG001
    """نسخ القالب ODS وتحويله لـ XLSX بـ LibreOffice مع كل التنسيق"""
    ods_name = f"Reort_Form_{report_type}.ods"
    ods_src = os.path.join(TEMPLATES_DIR, ods_name)

    if not os.path.exists(ods_src):
        print_err(f"القالب غير موجود: {ods_name}")
        log_file_op(
            f"قالب غير موجود: {ods_name}",
            operation="exists",
            file_path=ods_src,
            result="fail",
            error_reason="غير موجود",
            script_name="reports_generator",
            func_name="_convert_template",
        )
        return None

    # نسخ القالب لمجلد مؤقت ليحوله LibreOffice هناك
    tmp_dir = tempfile.mkdtemp(prefix="ht_report_")
    tmp_ods = os.path.join(tmp_dir, ods_name)
    shutil.copy2(ods_src, tmp_ods)
    log_file_op(
        f"نسخ قالب ODS: {ods_name}",
        operation="copy",
        file_path=ods_src,
        result="success",
        script_name="reports_generator",
        func_name="_convert_template",
    )

    # تحويل بـ LibreOffice - يحفظ التنسيق والألوان والتصميم بالكامل
    cmd = [SOFFICE, "--headless", "--convert-to", "xlsx", "--outdir", tmp_dir, tmp_ods]
    t0 = _time.time()
    try:
        result = subprocess.run(cmd, capture_output=True, timeout=30, check=False)  # noqa: S603
        dur = round(_time.time() - t0, 2)
        cmd_out = (result.stdout or b"").decode("utf-8", errors="replace")[:500]
        if result.returncode != 0:
            print_err("فشل التحويل عبر LibreOffice")
            log_file_op(
                "فشل LibreOffice",
                operation="subprocess",
                file_path=tmp_ods,
                result="fail",
                error_reason=f"returncode={result.returncode}",
                duration_seconds=dur,
                command_output=cmd_out,
                script_name="reports_generator",
                func_name="_convert_template",
            )
            shutil.rmtree(tmp_dir, ignore_errors=True)
            return None
        log_file_op(
            f"تحويل LibreOffice نجح ({dur}ث)",
            operation="subprocess",
            file_path=tmp_ods,
            result="success",
            duration_seconds=dur,
            command_output=cmd_out,
            script_name="reports_generator",
            func_name="_convert_template",
        )
    except FileNotFoundError:
        print_err("LibreOffice غير موجود!")
        log_exception(
            "LibreOffice غير موجود",
            script_name="reports_generator",
            func_name="_convert_template",
            context=SOFFICE,
        )
        shutil.rmtree(tmp_dir, ignore_errors=True)
        return None
    except subprocess.TimeoutExpired:
        print_err("انتهت مهلة التحويل!")
        log_exception(
            "مهلة LibreOffice انتهت",
            script_name="reports_generator",
            func_name="_convert_template",
            context="timeout=30s",
        )
        shutil.rmtree(tmp_dir, ignore_errors=True)
        return None

    # الملف المحول
    xlsx_name = ods_name.replace(".ods", ".xlsx")
    xlsx_path = os.path.join(tmp_dir, xlsx_name)
    if not os.path.exists(xlsx_path):
        print_err("الملف المحول غير موجود!")
        shutil.rmtree(tmp_dir, ignore_errors=True)
        return None

    return xlsx_path


# ──────────────────────────────────────────
# 5. تعبئة البيانات في الملف المحول
# ──────────────────────────────────────────


def _fill_data(xlsx_path, data, report_type) -> bool | None:
    """تعبئة القيم فقط - التنسيق لا يُمس"""
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        formula_cells = FORMULAS.get(report_type, set())
        written = 0

        for cell_ref, value in data.items():
            if cell_ref in formula_cells:
                continue
            if isinstance(value, (date, datetime)):
                ws[cell_ref].value = value.strftime("%d-%m-%Y")
                written += 1
            elif value is not None:
                ws[cell_ref].value = value
                written += 1

        # كشف نوبة بأصفار كلها — قد تعني مشكلة في البيانات المصدرية
        # الخلايا C3-C10 هي بيانات الأرقام لنوبة واحدة
        numeric_cells = [
            c for c in data if c.startswith("C") and c[1:].isdigit() and int(c[1:]) >= 3 and int(c[1:]) <= 21
        ]
        all_zero = all(data.get(c, 0) in (0, None, "") for c in numeric_cells)
        if all_zero and numeric_cells:
            log_quality(
                f"نوبة {report_type} بأصفار كلها",
                source_file=xlsx_path,
                issue_type="unknown_action",
                actual_value="كل القيم صفر",
                script_name="reports_generator",
                func_name="_fill_data",
            )

        # حقن معادلات غير موجودة في القالب
        for cell_ref, formula in INJECT_FORMULAS.get(report_type, {}).items():
            ws[cell_ref].value = formula

        wb.save(xlsx_path)
        wb.close()
        log_file_op(
            f"تعبئة {written} خلية في القالب",
            operation="write",
            file_path=xlsx_path,
            result="success",
            script_name="reports_generator",
            func_name="_fill_data",
        )
        return True
    except Exception as e:
        print_err(f"فشل تعبئة البيانات: {e}")
        log_exception(
            f"فشل تعبئة البيانات: {e}",
            exc=e,
            script_name="reports_generator",
            func_name="_fill_data",
        )
        return False


# ──────────────────────────────────────────
# 6. حفظ بالاسم النهائي
# ──────────────────────────────────────────


def _final_save(tmp_xlsx, source_file, report_type, date_value):
    """نقل الملف بجانب المصدر بالتسمية المطلوبة"""
    output_dir = os.path.dirname(source_file)

    if isinstance(date_value, (date, datetime)):
        date_str = date_value.strftime("%d-%m-%Y")
    elif isinstance(date_value, str):
        date_str = date_value
    else:
        date_str = datetime.now().strftime("%d-%m-%Y")

    filename = f"Reort_Form_{report_type}_{date_str}.xlsx"
    output = os.path.join(output_dir, filename)

    # ضمان عدم الكتابة فوق ملف موجود
    if os.path.exists(output):
        base, ext = os.path.splitext(output)
        counter = 1
        while os.path.exists(f"{base}_{counter}{ext}"):
            counter += 1
        output = f"{base}_{counter}{ext}"

    try:
        shutil.move(tmp_xlsx, output)
        out_size = None
        with contextlib.suppress(Exception):
            out_size = os.path.getsize(output)
        log_file_op(
            f"حفظ التقرير النهائي: {os.path.basename(output)}",
            operation="write",
            file_path=output,
            file_size_bytes=out_size,
            result="success",
            script_name="reports_generator",
            func_name="_final_save",
        )
        return output
    except Exception as e:
        print_err(f"فشل حفظ الملف: {e}")
        log_exception(
            f"فشل حفظ التقرير: {e}",
            exc=e,
            script_name="reports_generator",
            func_name="_final_save",
        )
        return None


# ──────────────────────────────────────────
# 7. عرض النتيجة
# ──────────────────────────────────────────


def _show_result(output_path) -> None:
    """رسالة نجاح مع خيارات"""
    print_ok("تم إنشاء التقرير بنجاح!")
    print(f"  {BOLD}{os.path.basename(output_path)}{RESET}")
    print(f"  {DIM}{output_path}{RESET}")
    print()
    print(f"    {CYAN}1{RESET}  فتح الملف")
    print(f"    {CYAN}2{RESET}  فتح المجلد")
    print(f"    {RED}0{RESET}  رجوع")

    choice = ask("اختر:")
    if choice == "1":
        try:
            os.startfile(output_path)
        except Exception as e:
            print_err(f"فشل الفتح: {e}")
    elif choice == "2":
        try:
            os.startfile(os.path.dirname(output_path))
        except Exception as e:
            print_err(f"فشل فتح المجلد: {e}")


# ──────────────────────────────────────────
# نقطة الدخول الرئيسية
# ──────────────────────────────────────────


def run_reports_menu(paths) -> None:
    """القائمة الرئيسية - تستقبل قائمة المسارات من مدير الملفات"""
    # فحص المتطلبات
    if not os.path.isdir(TEMPLATES_DIR):
        print_err("مجلد القوالب غير موجود!")
        print_err(f"المتوقع: {TEMPLATES_DIR}")
        return
    if not os.path.exists(SOFFICE):
        print_err("LibreOffice غير مثبت!")
        return
    if not paths:
        print_warn("لا توجد مسارات محفوظة")
        return

    t0 = _time.time()

    # 1. اختيار المسار
    selected = _select_path(paths)
    if not selected:
        return

    # 2. تصفح حتى اختيار الملف
    source_file = _browse_to_file(selected)
    if not source_file:
        return
    print_ok(f"الملف: {os.path.basename(source_file)}")

    # 3. اختيار نوع التقرير
    report_type = _select_report_type()
    if not report_type:
        return

    log_processing(
        f"بداية توليد تقرير {report_type}",
        operation_type="report_generation",
        input_file=source_file,
        script_name="reports_generator",
        func_name="run_reports_menu",
    )

    # 4. قراءة البيانات من المصدر
    print(f"  {DIM}جاري قراءة البيانات...{RESET}")
    data = _read_source(source_file, MAPPINGS[report_type])
    if not data:
        log_processing(
            "فشل قراءة البيانات",
            operation_type="report_generation",
            input_file=source_file,
            result="fail",
            duration_seconds=round(_time.time() - t0, 2),
            script_name="reports_generator",
            func_name="run_reports_menu",
        )
        return

    # تسجيل التاريخ المرجعي والطابع
    ref_date = data.get("A2", "")
    log_processing(
        f"التاريخ المرجعي: {ref_date}",
        operation_type="report_generation",
        input_file=source_file,
        script_name="reports_generator",
        func_name="run_reports_menu",
    )

    # 5. تحويل القالب بـ LibreOffice (تنسيق كامل)
    print(f"  {DIM}جاري تجهيز القالب...{RESET}")
    tmp_xlsx = _convert_template(report_type, tempfile.gettempdir())
    if not tmp_xlsx:
        return

    # 6. تعبئة القيم فقط
    print(f"  {DIM}جاري تعبئة البيانات...{RESET}")
    if not _fill_data(tmp_xlsx, data, report_type):
        return

    # 7. حفظ بجانب الملف المصدر
    date_value = data.get("A2", datetime.now().date())
    output = _final_save(tmp_xlsx, source_file, report_type, date_value)
    if not output:
        return

    # تنظيف المجلد المؤقت
    tmp_dir = os.path.dirname(tmp_xlsx)
    shutil.rmtree(tmp_dir, ignore_errors=True)

    duration = round(_time.time() - t0, 2)
    log_processing(
        f"نجاح توليد تقرير {report_type}",
        operation_type="report_generation",
        input_file=source_file,
        output_file=output,
        duration_seconds=duration,
        result="success",
        script_name="reports_generator",
        func_name="run_reports_menu",
    )

    # 8. عرض النتيجة
    _show_result(output)
