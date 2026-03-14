"""
مزامنة مجلد Record_2026 مع قاعدة البيانات الجديدة ht_sc_new.db.

يقرأ ملفات الإكسل اليومية ويخزّنها في جدول record_2026.
يراقب المجلد تلقائياً لالتقاط أي ملف جديد أو تعديل.

تشغيل مباشر: python db/sync_record_2026.py
"""

import logging
import os
import re
import sqlite3
import time
from pathlib import Path

import openpyxl
from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer

from constants import MAX_TRIPS, TRIP_FIELDS
from helpers import build_record_2026_insert_sql, safe_int, safe_str

# ═══════════════════════════════════════════════════════
# الإعدادات
# ═══════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("sync_record_2026")

# مسارات
DB_DIR = Path(__file__).parent
DB_PATH = DB_DIR / "ht_sc_new.db"
RECORD_DIR = Path(__file__).parent.parent / "Record_2026"

# عدد الأعمدة الثابتة قبل حقول الرحلات
# رمز النوبة [0] + اسم المرسل [1] + وقت الرسالة [2]
FIXED_COLS = 3

# حدود النوبات — بداية كل نوبة (ساعة, دقيقة)
SHIFT_BOUNDARIES = {
    "A": (5, 45),    # 05:45 — 13:44
    "B": (13, 45),   # 13:45 — 21:44
    "C": (21, 45),   # 21:45 — 05:44
}


def get_current_shift() -> tuple[str, str]:
    """تحديد النوبة الحالية والتاريخ بناءً على الساعة الفعلية.

    النوبة C تمتد من 21:45 حتى 05:44 — التاريخ يتبع بداية النوبة.
    إرجاع: (رمز_النوبة, التاريخ بصيغة DD-MM-YYYY)
    """
    from datetime import datetime, timedelta

    now = datetime.now()
    hour, minute = now.hour, now.minute
    current = hour * 60 + minute

    a_start = 5 * 60 + 45    # 05:45 = 345
    b_start = 13 * 60 + 45   # 13:45 = 825
    c_start = 21 * 60 + 45   # 21:45 = 1305

    if a_start <= current < b_start:
        shift = "A"
        shift_date = now
    elif b_start <= current < c_start:
        shift = "B"
        shift_date = now
    else:
        shift = "C"
        # بعد منتصف الليل وقبل 05:45 — التاريخ يتبع أمس
        if current < a_start:
            shift_date = now - timedelta(days=1)
        else:
            shift_date = now

    return shift, shift_date.strftime("%d-%m-%Y")


# ═══════════════════════════════════════════════════════
# 1. إنشاء الجدول والفهارس
# ═══════════════════════════════════════════════════════


def get_connection(db_path: str | Path | None = None) -> sqlite3.Connection:
    """إنشاء اتصال بالقاعدة الجديدة مع تفعيل المفاتيح الأجنبية."""
    path = str(db_path or DB_PATH)
    conn = sqlite3.connect(path)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    return conn


def create_table(conn: sqlite3.Connection) -> None:
    """إنشاء جدول record_2026 مع أعمدة الرحلات المرقّمة."""
    # بناء أعمدة الرحلات ديناميكياً
    trip_cols = []
    for n in range(1, MAX_TRIPS + 1):
        trip_cols.append(f"""
            flight_number_{n}   TEXT,
            departure_time_{n}  TEXT,
            passenger_count_{n} INTEGER DEFAULT 0,
            destination_{n}     TEXT,
            visa_type_{n}       TEXT,
            campaign_name_{n}   TEXT,
            status_{n}          TEXT,
            dispatch_{n}        TEXT,
            inspection_{n}      TEXT""")

    sql = f"""
        CREATE TABLE IF NOT EXISTS record_2026 (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            folder_date     TEXT NOT NULL,
            filename        TEXT,
            shift_code      TEXT NOT NULL,
            row_num         INTEGER,
            sender_name     TEXT,
            msg_time        TEXT,
            trip_count      INTEGER DEFAULT 1,
            {",".join(trip_cols)},
            synced_at       TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(filename, row_num)
        )
    """
    conn.execute(sql)
    conn.commit()


def create_indexes(conn: sqlite3.Connection) -> None:
    """إنشاء الفهارس لتسريع الاستعلامات."""
    indexes = [
        ("idx_record2026_date", "record_2026(folder_date)"),
        ("idx_record2026_shift", "record_2026(shift_code)"),
        ("idx_record2026_flight", "record_2026(flight_number_1)"),
        ("idx_record2026_sender", "record_2026(sender_name)"),
    ]
    for name, target in indexes:
        conn.execute(
            f"CREATE INDEX IF NOT EXISTS {name} ON {target}"
        )
    conn.commit()


def create_reference_tables(conn: sqlite3.Connection) -> None:
    """إنشاء الجداول المرجعية للتقارير — بدون قيود مفاتيح أجنبية."""
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS shifts (
            code       TEXT PRIMARY KEY,
            name_ar    TEXT NOT NULL,
            start_time TEXT NOT NULL,
            end_time   TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS destinations (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT    NOT NULL UNIQUE
        );

        CREATE TABLE IF NOT EXISTS airlines (
            code TEXT PRIMARY KEY,
            name TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS visa_types (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT    NOT NULL UNIQUE
        );

        CREATE TABLE IF NOT EXISTS flight_status (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT    NOT NULL UNIQUE
        );

        CREATE TABLE IF NOT EXISTS dispatching (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT    NOT NULL UNIQUE
        );

        CREATE TABLE IF NOT EXISTS actions_taken (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            description TEXT    NOT NULL UNIQUE
        );

        CREATE TABLE IF NOT EXISTS jadco_supervisors (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT    NOT NULL UNIQUE
        );

        CREATE TABLE IF NOT EXISTS agent_supervisors (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT    NOT NULL UNIQUE
        );

        CREATE TABLE IF NOT EXISTS travel_halls (
            code TEXT PRIMARY KEY,
            name TEXT NOT NULL UNIQUE
        );

        CREATE TABLE IF NOT EXISTS passenger_status (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT    NOT NULL UNIQUE
        );

        CREATE TABLE IF NOT EXISTS ground_services (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT    NOT NULL UNIQUE
        );
    """)


def seed_reference_data(conn: sqlite3.Connection) -> None:
    """تعبئة الجداول المرجعية من قوائم المشروع."""
    import sys
    lists_dir = str(Path(__file__).parent.parent / "docs" / "List")
    if lists_dir not in sys.path:
        sys.path.insert(0, lists_dir)

    from lists import (
        actions_taken, agent_supervisors, airline_codes,
        airline_names, destinations, dispatching,
        flight_status, ground_services, jadco_supervisors,
        passenger_status, travel_hall_codes, travel_halls,
        visa_types,
    )

    # النوبات الثلاث
    conn.executemany(
        "INSERT OR IGNORE INTO shifts (code, name_ar, start_time, end_time) "
        "VALUES (?, ?, ?, ?)",
        [
            ("A", "صباح", "05:45:00", "13:44:59"),
            ("B", "ظهر", "13:45:00", "21:44:59"),
            ("C", "مساء", "21:45:00", "05:44:59"),
        ],
    )

    # جداول العمود الواحد
    single_col = {
        "destinations": destinations,
        "dispatching": dispatching,
        "flight_status": flight_status,
        "visa_types": visa_types,
        "jadco_supervisors": jadco_supervisors,
        "agent_supervisors": agent_supervisors,
        "passenger_status": passenger_status,
        "ground_services": ground_services,
    }
    for table, data in single_col.items():
        col = "description" if table == "actions_taken" else "name"
        conn.executemany(
            f"INSERT OR IGNORE INTO {table} ({col}) VALUES (?)",
            [(v,) for v in data],
        )

    # الإجراءات المتخذة
    conn.executemany(
        "INSERT OR IGNORE INTO actions_taken (description) VALUES (?)",
        [(v,) for v in actions_taken],
    )

    # الناقلات الجوية — عمودان
    conn.executemany(
        "INSERT OR IGNORE INTO airlines (code, name) VALUES (?, ?)",
        list(zip(airline_codes, airline_names)),
    )

    # صالات السفر — عمودان
    conn.executemany(
        "INSERT OR IGNORE INTO travel_halls (code, name) VALUES (?, ?)",
        list(zip(travel_hall_codes, travel_halls)),
    )

    conn.commit()


def _create_shift_table(conn: sqlite3.Connection, shift: str) -> None:
    """إنشاء جدول إحصائي لنوبة واحدة."""
    table = f"shift_{shift.lower()}_stats"
    conn.execute(f"""
        CREATE TABLE IF NOT EXISTS {table} (
            id                     INTEGER PRIMARY KEY AUTOINCREMENT,
            folder_date            TEXT NOT NULL UNIQUE,
            dep_buses              INTEGER DEFAULT 0,
            umrah_buses            INTEGER DEFAULT 0,
            tourist_buses          INTEGER DEFAULT 0,
            dep_pax                INTEGER DEFAULT 0,
            umrah_pax              INTEGER DEFAULT 0,
            tourist_pax            INTEGER DEFAULT 0,
            arrival_buses          INTEGER DEFAULT 0,
            bus_accidents          INTEGER DEFAULT 0,
            car_accidents          INTEGER DEFAULT 0,
            early                  INTEGER DEFAULT 0,
            late                   INTEGER DEFAULT 0,
            wrong                  INTEGER DEFAULT 0,
            mix                    INTEGER DEFAULT 0,
            minutes                INTEGER DEFAULT 0,
            stumbler               INTEGER,
            operational_letters    INTEGER DEFAULT 0,
            weather_conditions     INTEGER DEFAULT 0,
            operational_reasons    INTEGER DEFAULT 0,
            emergency_maintenance  INTEGER DEFAULT 0,
            other_notes            INTEGER DEFAULT 0,
            updated_at             TEXT DEFAULT (datetime('now','localtime'))
        )
    """)


def create_shift_tables(conn: sqlite3.Connection) -> None:
    """إنشاء الجداول الثلاثة للنوبات + الجدول اليومي."""
    for shift in ("A", "B", "C"):
        _create_shift_table(conn, shift)

    # جدول الإحصائية اليومية — مجموع النوبات الثلاث
    conn.execute("""
        CREATE TABLE IF NOT EXISTS daily_stats (
            id                     INTEGER PRIMARY KEY AUTOINCREMENT,
            folder_date            TEXT NOT NULL UNIQUE,
            dep_buses              INTEGER DEFAULT 0,
            umrah_buses            INTEGER DEFAULT 0,
            tourist_buses          INTEGER DEFAULT 0,
            dep_pax                INTEGER DEFAULT 0,
            umrah_pax              INTEGER DEFAULT 0,
            tourist_pax            INTEGER DEFAULT 0,
            arrival_buses          INTEGER DEFAULT 0,
            bus_accidents          INTEGER DEFAULT 0,
            car_accidents          INTEGER DEFAULT 0,
            early                  INTEGER DEFAULT 0,
            late                   INTEGER DEFAULT 0,
            wrong                  INTEGER DEFAULT 0,
            mix                    INTEGER DEFAULT 0,
            stumbler               INTEGER,
            operational_letters    INTEGER DEFAULT 0,
            weather_conditions     INTEGER DEFAULT 0,
            operational_reasons    INTEGER DEFAULT 0,
            emergency_maintenance  INTEGER DEFAULT 0,
            other_notes            INTEGER DEFAULT 0,
            total                  INTEGER DEFAULT 0,
            updated_at             TEXT DEFAULT (datetime('now','localtime'))
        )
    """)
    conn.commit()


def refresh_shift_stats(
    conn: sqlite3.Connection,
    shift: str | None = None,
) -> None:
    """حساب الإحصائيات وتعبئة جدول النوبة من record_2026."""
    shifts = [shift] if shift else ["A", "B", "C"]

    for s in shifts:
        table = f"shift_{s.lower()}_stats"

        # شروط فيزا العمرة — عمره وعمرة
        umrah_visa = " OR ".join(
            f"visa_type_{n} LIKE '%عمر%'"
            for n in range(1, MAX_TRIPS + 1)
        )

        # شروط فيزا السياحة — سياحة وسياحه
        tourist_visa = " OR ".join(
            f"visa_type_{n} LIKE '%سياح%'"
            for n in range(1, MAX_TRIPS + 1)
        )

        # مجموع الركاب الكلي (1-10)
        total_pax = " + ".join(
            f"COALESCE(passenger_count_{n}, 0)"
            for n in range(1, MAX_TRIPS + 1)
        )

        # ركاب العمرة فقط
        umrah_pax = " + ".join(
            f"CASE WHEN visa_type_{n} LIKE '%عمر%'"
            f" THEN COALESCE(passenger_count_{n}, 0) ELSE 0 END"
            for n in range(1, MAX_TRIPS + 1)
        )

        # ركاب السياحة فقط
        tourist_pax = " + ".join(
            f"CASE WHEN visa_type_{n} LIKE '%سياح%'"
            f" THEN COALESCE(passenger_count_{n}, 0) ELSE 0 END"
            for n in range(1, MAX_TRIPS + 1)
        )

        # التفويج — من كل أعمدة التفويج (1-10)
        early = " + ".join(
            f"CASE WHEN dispatch_{n} LIKE '%مبكر%'"
            f" THEN 1 ELSE 0 END"
            for n in range(1, MAX_TRIPS + 1)
        )

        late = " + ".join(
            f"CASE WHEN dispatch_{n} LIKE '%تأخر%'"
            f" OR dispatch_{n} LIKE '%تاخر%'"
            f" THEN 1 ELSE 0 END"
            for n in range(1, MAX_TRIPS + 1)
        )

        wrong = " + ".join(
            f"CASE WHEN dispatch_{n} LIKE '%خاط%'"
            f" THEN 1 ELSE 0 END"
            for n in range(1, MAX_TRIPS + 1)
        )

        mix = " + ".join(
            f"CASE WHEN dispatch_{n} LIKE '%مشترك%'"
            f" THEN 1 ELSE 0 END"
            for n in range(1, MAX_TRIPS + 1)
        )

        conn.execute(f"DELETE FROM {table}")
        conn.execute(f"""
            INSERT INTO {table} (
                folder_date, dep_buses, umrah_buses, tourist_buses,
                dep_pax, umrah_pax, tourist_pax,
                arrival_buses, bus_accidents, car_accidents,
                early, late, wrong, mix, minutes,
                stumbler,
                operational_letters, weather_conditions,
                operational_reasons, emergency_maintenance,
                other_notes
            )
            SELECT
                folder_date,
                COUNT(*),
                SUM(CASE WHEN {umrah_visa}
                    THEN 1 ELSE 0 END),
                SUM(CASE WHEN {tourist_visa}
                    THEN 1 ELSE 0 END),
                SUM({total_pax}),
                SUM({umrah_pax}),
                SUM({tourist_pax}),
                0, 0, 0,
                SUM({early}),
                SUM({late}),
                SUM({wrong}),
                SUM({mix}),
                SUM({early}) + SUM({late})
                    + SUM({wrong}) + SUM({mix}),
                NULL,
                0, 0, 0, 0, 0
            FROM record_2026
            WHERE shift_code = '{s}'
            GROUP BY folder_date
        """)

        logger.info(
            f"تحديث إحصائيات نوبة {s}: "
            f"{conn.execute(f'SELECT COUNT(*) FROM {table}').fetchone()[0]} يوم"
        )

    conn.commit()


def refresh_daily_stats(conn: sqlite3.Connection) -> None:
    """تعبئة الجدول اليومي بجمع النوبات الثلاث لكل يوم."""
    conn.execute("DELETE FROM daily_stats")
    conn.execute("""
        INSERT INTO daily_stats (
            folder_date, dep_buses, umrah_buses, tourist_buses,
            dep_pax, umrah_pax, tourist_pax,
            arrival_buses, bus_accidents, car_accidents,
            early, late, wrong, mix,
            stumbler,
            operational_letters, weather_conditions,
            operational_reasons, emergency_maintenance,
            other_notes, total
        )
        SELECT
            folder_date,
            SUM(dep_buses),
            SUM(umrah_buses),
            SUM(tourist_buses),
            SUM(dep_pax),
            SUM(umrah_pax),
            SUM(tourist_pax),
            SUM(arrival_buses),
            SUM(bus_accidents),
            SUM(car_accidents),
            SUM(early),
            SUM(late),
            SUM(wrong),
            SUM(mix),
            NULL,
            SUM(operational_letters),
            SUM(weather_conditions),
            SUM(operational_reasons),
            SUM(emergency_maintenance),
            SUM(other_notes),
            SUM(dep_buses) + SUM(umrah_buses) + SUM(tourist_buses)
            + SUM(dep_pax) + SUM(umrah_pax) + SUM(tourist_pax)
            + SUM(arrival_buses) + SUM(bus_accidents) + SUM(car_accidents)
            + SUM(early) + SUM(late) + SUM(wrong) + SUM(mix)
        FROM (
            SELECT * FROM shift_a_stats
            UNION ALL
            SELECT * FROM shift_b_stats
            UNION ALL
            SELECT * FROM shift_c_stats
        )
        GROUP BY folder_date
    """)
    logger.info(
        f"تحديث الإحصائية اليومية: "
        f"{conn.execute('SELECT COUNT(*) FROM daily_stats').fetchone()[0]} يوم"
    )
    conn.commit()


def init_db(db_path: str | Path | None = None) -> sqlite3.Connection:
    """تهيئة القاعدة: اتصال + جداول + فهارس + بيانات مرجعية + إحصائيات."""
    conn = get_connection(db_path)
    create_table(conn)
    create_indexes(conn)
    create_reference_tables(conn)
    seed_reference_data(conn)
    create_shift_tables(conn)
    return conn


# ═══════════════════════════════════════════════════════
# 2. أدوات التحويل
# ═══════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════
# 3. قراءة ملف إكسل واحد
# ═══════════════════════════════════════════════════════

# جملة الإدراج — تُبنى مرة واحدة عند تحميل الوحدة
_INSERT_SQL = build_record_2026_insert_sql()


def _extract_shift_letter(raw_code: str) -> str:
    """استخراج حرف النوبة من الرمز الكامل (00A# → A)."""
    raw = raw_code.strip().upper()
    # البحث عن A أو B أو C في أي موضع
    for ch in raw:
        if ch in ("A", "B", "C"):
            return ch
    return raw


def parse_file_info(file_path: Path) -> dict | None:
    """استخراج معلومات النوبة من اسم ملف الإكسل."""
    name = file_path.name
    if not name.endswith(".xlsx"):
        return None
    # تجاهل الملفات المؤقتة والنسخ الاحتياطية
    if name.endswith((".tmp", ".bak")):
        return None
    # صيغة: D(D)-MM-YYYY-Record-X.xlsx
    match = re.match(
        r"^\d{1,2}-\d{2}-\d{4}-Record-([ABC])\.xlsx$", name
    )
    if not match:
        return None
    return {
        "filename": name,
        "shift_code": match.group(1),
    }


def read_excel_file(
    file_path: Path,
    folder_date: str,
) -> list[tuple]:
    """قراءة ملف إكسل وإرجاع صفوف جاهزة للإدراج."""
    file_info = parse_file_info(file_path)
    if not file_info:
        return []

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
    except Exception as e:
        logger.error(f"فشل فتح {file_path.name}: {e}")
        return []

    ws = wb.active
    if ws is None or ws.max_row is None or ws.max_row < 2:
        wb.close()
        return []

    rows_data = []
    col_count = ws.max_column or 0

    # كشف عمود المعرّف 'م' — إزاحة الأعمدة
    first_header = safe_str(ws.cell(1, 1).value)
    off = 1 if first_header == "م" else 0

    for row_idx in range(2, ws.max_row + 1):
        cells = []
        for col in range(1, col_count + 1):
            cells.append(ws.cell(row_idx, col).value)

        # تجاهل الصفوف الفارغة
        if not any(c is not None for c in cells):
            continue

        # الأعمدة الثابتة — مع إزاحة عمود المعرّف
        raw_shift = safe_str(cells[off]) if len(cells) > off else ""
        shift_code = _extract_shift_letter(raw_shift)
        sender_name = (
            safe_str(cells[off + 1]) if len(cells) > off + 1 else ""
        )
        msg_time = (
            safe_str(cells[off + 2]) if len(cells) > off + 2 else ""
        )

        # حساب عدد الرحلات الفعلية
        trip_count = 0
        for t in range(MAX_TRIPS):
            base = FIXED_COLS + off + (t * 9)
            trip_slice = cells[base:base + 9]
            if any(c is not None for c in trip_slice):
                trip_count = t + 1

        if trip_count == 0:
            trip_count = 1

        # بناء قيم الرحلات
        trip_values = []
        for t in range(MAX_TRIPS):
            base = FIXED_COLS + off + (t * 9)
            for i, field in enumerate(TRIP_FIELDS):
                idx = base + i
                if (
                    t < trip_count
                    and idx < len(cells)
                    and cells[idx] is not None
                ):
                    if field == "passenger_count":
                        trip_values.append(safe_int(cells[idx]))
                    else:
                        trip_values.append(safe_str(cells[idx]))
                else:
                    if field == "passenger_count":
                        trip_values.append(0)
                    else:
                        trip_values.append(None)

        # القيم الثابتة
        fixed_values = [
            folder_date,
            file_info["filename"],
            shift_code,
            row_idx - 1,  # رقم الصف (بدون العنوان)
            sender_name,
            msg_time,
            trip_count,
        ]

        rows_data.append(tuple(fixed_values + trip_values))

    wb.close()
    return rows_data


# ═══════════════════════════════════════════════════════
# 4. المزامنة
# ═══════════════════════════════════════════════════════


def sync_single_file(
    conn: sqlite3.Connection,
    file_path: Path,
    folder_date: str,
) -> int:
    """مزامنة ملف إكسل واحد — إرجاع عدد الصفوف المُدرجة."""
    rows = read_excel_file(file_path, folder_date)
    if not rows:
        return 0

    inserted = 0
    # (رقم_الصف_في_الإكسل, record_2026.id) — لكتابة المعرّف رجوعاً
    id_updates: list[tuple[int, int]] = []

    for row in rows:
        try:
            cursor = conn.execute(_INSERT_SQL, row)
            inserted += 1
            # جمع المعرّف للصفوف الجديدة فقط
            if cursor.rowcount > 0 and cursor.lastrowid is not None:
                id_updates.append((int(row[3]), int(cursor.lastrowid)))
        except sqlite3.IntegrityError:
            # الصف موجود مسبقاً — تخطّي
            pass

    conn.commit()

    # كتابة record_2026.id في عمود 'م' بالإكسل
    if id_updates:
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            if ws is not None:
                first_header = safe_str(ws.cell(1, 1).value)
                if first_header == "م":
                    for row_num, record_id in id_updates:
                        # row_num = رقم الصف بدون العنوان — صف الإكسل = row_num + 1
                        ws.cell(row_num + 1, 1).value = record_id  # type: ignore[union-attr]
                    wb.save(file_path)
            wb.close()
        except Exception as e:
            logger.error(f"خطأ كتابة المعرّف في {file_path.name}: {e}")

    if inserted > 0:
        logger.info(f"{file_path.name}: {inserted} صف جديد")
        # تحديث إحصائيات النوبة + اليومي بعد كل ملف جديد
        file_info = parse_file_info(file_path)
        if file_info:
            refresh_shift_stats(conn, file_info["shift_code"])
            refresh_daily_stats(conn)
    return inserted


def full_sync(conn: sqlite3.Connection) -> int:
    """مزامنة كاملة لجميع مجلدات الأشهر والأيام."""
    if not RECORD_DIR.exists():
        logger.warning(f"المجلد غير موجود: {RECORD_DIR}")
        return 0

    total = 0
    # مسح مجلدات الأشهر (03-2026, 04-2026, ...)
    for month_dir in sorted(RECORD_DIR.iterdir()):
        if not month_dir.is_dir():
            continue
        if not re.match(r"^\d{2}-\d{4}$", month_dir.name):
            continue

        # مسح مجلدات الأيام (11-03-2026, 12-03-2026, ...)
        for day_dir in sorted(month_dir.iterdir()):
            if not day_dir.is_dir():
                continue
            if not re.match(r"^\d{2}-\d{2}-\d{4}$", day_dir.name):
                continue

            folder_date = day_dir.name

            # مسح ملفات الإكسل
            for xlsx_file in sorted(day_dir.glob("*.xlsx")):
                if xlsx_file.name.endswith((".tmp", ".bak")):
                    continue
                if parse_file_info(xlsx_file) is None:
                    continue
                total += sync_single_file(conn, xlsx_file, folder_date)

    # تحديث إحصائيات جميع النوبات + اليومي بعد المزامنة الكاملة
    if total > 0:
        refresh_shift_stats(conn)
        refresh_daily_stats(conn)

    logger.info(f"المزامنة الكاملة: {total} صف جديد")
    return total


# ═══════════════════════════════════════════════════════
# 5. المراقبة التلقائية
# ═══════════════════════════════════════════════════════


class RecordFileHandler(FileSystemEventHandler):
    """مراقب أحداث الملفات — يزامن عند إنشاء أو تعديل ملف إكسل."""

    def __init__(self, db_path: str | Path):
        """تهيئة المراقب مع مسار القاعدة."""
        super().__init__()
        self.db_path = db_path

    def _handle(self, file_path: Path) -> None:
        """معالجة حدث ملف واحد."""
        if not file_path.name.endswith(".xlsx"):
            return
        if file_path.name.endswith((".tmp", ".bak")):
            return
        if parse_file_info(file_path) is None:
            return

        # استخراج تاريخ المجلد
        parent = file_path.parent
        if not re.match(r"^\d{2}-\d{2}-\d{4}$", parent.name):
            return

        # انتظار قصير لاكتمال الكتابة
        time.sleep(0.5)

        try:
            conn = get_connection(self.db_path)
            sync_single_file(conn, file_path, parent.name)
            conn.close()
        except Exception as e:
            logger.error(f"خطأ مزامنة {file_path.name}: {e}")

    def on_created(self, event):
        """عند إنشاء ملف جديد."""
        if not event.is_directory:
            self._handle(Path(os.fsdecode(event.src_path)))

    def on_modified(self, event):
        """عند تعديل ملف موجود."""
        if not event.is_directory:
            self._handle(Path(os.fsdecode(event.src_path)))

    def on_moved(self, event):
        """عند نقل أو إعادة تسمية ملف."""
        if not event.is_directory and hasattr(event, "dest_path"):
            self._handle(Path(os.fsdecode(event.dest_path)))


def start_watcher(db_path: str | Path | None = None) -> None:
    """تشغيل المراقب التلقائي مع مزامنة كاملة أولية."""
    path = db_path or DB_PATH

    # تحديد النوبة الحالية
    shift_code, shift_date = get_current_shift()
    logger.info(f"النوبة الحالية: {shift_code} — التاريخ: {shift_date}")

    # تهيئة القاعدة + مزامنة كاملة
    conn = init_db(path)
    full_sync(conn)

    # عرض إحصائية النوبة الحالية
    table = f"shift_{shift_code.lower()}_stats"
    row = conn.execute(
        f"SELECT * FROM {table} WHERE folder_date = ?",
        (shift_date,),
    ).fetchone()

    if row:
        logger.info(
            f"إحصائية نوبة {shift_code} ليوم {shift_date}: "
            f"حافلات={row['dep_buses']} "
            f"عمرة={row['umrah_buses']} "
            f"سياحة={row['tourist_buses']} "
            f"ركاب={row['dep_pax']} "
            f"مبكر={row['early']} "
            f"متأخر={row['late']} "
            f"خاطئ={row['wrong']} "
            f"مشترك={row['mix']}"
        )
    else:
        logger.info(
            f"لا توجد بيانات لنوبة {shift_code} ليوم {shift_date}"
        )

    conn.close()

    # تشغيل المراقب
    handler = RecordFileHandler(path)
    observer = Observer()
    observer.schedule(handler, str(RECORD_DIR), recursive=True)
    observer.start()
    logger.info(f"المراقب يعمل على: {RECORD_DIR}")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
        logger.info("تم إيقاف المراقب")
    observer.join()


# ═══════════════════════════════════════════════════════
# التشغيل المباشر
# ═══════════════════════════════════════════════════════

if __name__ == "__main__":
    # تهيئة + مزامنة + عرض النتائج
    conn = init_db()
    total = full_sync(conn)

    # عرض إحصائيات
    count = conn.execute(
        "SELECT COUNT(*) FROM record_2026"
    ).fetchone()[0]
    days = conn.execute(
        "SELECT COUNT(DISTINCT folder_date) FROM record_2026"
    ).fetchone()[0]

    print(f"\nإجمالي الصفوف: {count}")
    print(f"عدد الأيام: {days}")

    # تفصيل حسب اليوم والنوبة
    rows = conn.execute(
        "SELECT folder_date, shift_code, COUNT(*) as cnt "
        "FROM record_2026 "
        "GROUP BY folder_date, shift_code "
        "ORDER BY folder_date, shift_code"
    ).fetchall()

    for r in rows:
        print(f"  {r['folder_date']} | {r['shift_code']} | {r['cnt']} صف")

    conn.close()
