"""
استيراد بيانات التسجيل اليومي (RECORD) إلى قاعدة البيانات.

يقرأ ملف إكسل ويحفظ كل صف كسجل في جدول records
بالهيكل الجديد: 8 أعمدة ثابتة + 90 عمود رحلات (9 حقول × 10 رحلات)

الاستخدام:
    from db.import_record import import_record
    stats = import_record("path/to/record.xlsx", date="01-01-2026", shift="A")
"""

import os
import sys


_PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_DIR not in sys.path:
    sys.path.insert(0, _PROJECT_DIR)
_SCRIPTS_DIR = os.path.join(_PROJECT_DIR, "scripts")
if _SCRIPTS_DIR not in sys.path:
    sys.path.insert(0, _SCRIPTS_DIR)

from db.constants import MAX_TRIPS, TRIP_FIELDS
from db.helpers import build_records_insert_sql, safe_int, safe_str
from db.schema import DB_PATH, get_connection
from logger_config import log_db_op, log_exception

# === أنماط مطابقة أعمدة التسجيل ===
# تطابق جزئي وغير حساس لحالة الأحرف — أول تطابق يُعتمد
_COL_PATTERNS = {
    "sender_name": ["المرسل", "اسم المرسل", "sender", "اسم"],
    "msg_time": ["وقت الرسالة", "وقت الوصول", "msg time", "وقت"],
    "flight_number": ["رقم الرحلة", "flight no", "flight number", "رقم رحلة"],
    "departure_time": ["وقت الاقلاع", "std", "departure", "وقت اقلاع"],
    "passenger_count": ["عدد الركاب", "pax", "passenger", "عدد ركاب"],
    "destination": ["الوجهة", "destination", "dest", "وجهة"],
    "visa_type": ["الفيزا", "visa type", "visa", "فيزا"],
    "campaign_name": ["اسم الحملة", "company", "campaign", "حملة"],
    "status": ["الحالة", "flight status", "حالة الرحلة"],
    "dispatch": ["التفويج", "تفويج"],
    "inspection": ["الكشف", "كشف"],
}


def _map_columns(headers_dict):
    """بناء خريطة: اسم الحقل ← رقم العمود في الإكسل."""
    col_map = {}
    for field, patterns in _COL_PATTERNS.items():
        for col_idx, header_name in headers_dict.items():
            header_low = str(header_name).lower().strip()
            for pat in patterns:
                if pat.lower() in header_low:
                    col_map[field] = col_idx
                    break
            if field in col_map:
                break
    return col_map


# جملة الإدراج — تُبنى مرة واحدة
_INSERT_SQL = build_records_insert_sql()


def import_record(record_path, date, shift, db_path=None, file_id=None):
    """
    استيراد ملف تسجيل يومي واحد إلى قاعدة البيانات.

    المدخلات:
        record_path: مسار ملف الإكسل
        date: التاريخ بأي صيغة (يُحفظ كما هو)
        shift: رمز النوبة (A/B/C)
        db_path: مسار القاعدة (اختياري)
        file_id: معرّف الملف في جدول excel_files (اختياري)

    المخرجات:
        قاموس: rows, total_pax, trips, mapped_cols, unmapped_cols
    """
    if not os.path.exists(record_path):
        raise FileNotFoundError(f"ملف التسجيل غير موجود: {record_path}")

    import openpyxl

    wb = openpyxl.load_workbook(record_path, read_only=True, data_only=True)

    try:
        # البحث عن ورقة Record
        sheet = None
        for name in wb.sheetnames:
            if name.lower() == "record":
                sheet = wb[name]
                break
        if sheet is None:
            sheet = wb.active

        # قراءة الرؤوس من سطر 8
        headers = {}
        for col_idx, cell in enumerate(sheet[8], 1):
            if cell.value:
                headers[col_idx] = str(cell.value).strip()

        if not headers:
            raise ValueError("لا توجد رؤوس في سطر 8 — هل الملف بالتنسيق الصحيح؟")

        # بناء خريطة الأعمدة
        col_map = _map_columns(headers)

        # قراءة البيانات من سطر 9
        raw_rows = []
        for row in sheet.iter_rows(min_row=9, values_only=False):
            vals = [c.value for c in row]
            if all(v is None for v in vals):
                continue
            row_data = {"_row_num": row[0].row}
            for col_idx in headers:
                if col_idx <= len(row):
                    row_data[col_idx] = row[col_idx - 1].value
            raw_rows.append(row_data)

    finally:
        wb.close()

    if not raw_rows:
        return {
            "rows": 0,
            "total_pax": 0,
            "trips": 0,
            "mapped_cols": list(col_map.keys()),
            "unmapped_cols": [],
        }

    # الأعمدة غير المكتشفة — للتوثيق
    trip_fields_set = set(TRIP_FIELDS)
    mapped_trip_fields = set(col_map.keys()) & trip_fields_set
    unmapped = list(trip_fields_set - mapped_trip_fields)

    # الإدراج في القاعدة
    conn = get_connection(db_path)
    cur = conn.cursor()
    total_pax = 0
    inserted = 0

    try:
        for raw in raw_rows:
            row_num = raw["_row_num"]

            # استخراج رقم الرحلة — تخطي الصفوف بدون رحلة
            flight = safe_str(raw.get(col_map.get("flight_number")))
            if not flight:
                continue

            # استخراج المرسل والوقت إن وُجدا
            sender_name = safe_str(raw.get(col_map.get("sender_name")))
            msg_time = safe_str(raw.get(col_map.get("msg_time")))

            # استخراج حقول الرحلة الواحدة
            dep_time = safe_str(raw.get(col_map.get("departure_time")))
            pax = safe_int(raw.get(col_map.get("passenger_count")))
            dest = safe_str(raw.get(col_map.get("destination")))
            visa = safe_str(raw.get(col_map.get("visa_type")))
            campaign = safe_str(raw.get(col_map.get("campaign_name")))
            status_val = safe_str(raw.get(col_map.get("status")))
            dispatch_val = safe_str(raw.get(col_map.get("dispatch")))
            inspection_val = safe_str(raw.get(col_map.get("inspection")))

            total_pax += pax

            # الأعمدة الثابتة (7 قيم)
            fixed_values = [
                date,
                shift,
                sender_name,
                msg_time,
                row_num,
                1,  # trip_count — صف واحد = رحلة واحدة
                file_id,
            ]

            # قيم الرحلات: الرحلة الأولى فقط + باقي الخانات فارغة
            trip_values = [
                flight,
                dep_time,
                pax,
                dest,
                visa,
                campaign,
                status_val,
                dispatch_val,
                inspection_val,
            ]
            # تعبئة الرحلات 2-10 بقيم فارغة (9 حقول × 9 رحلات)
            for _ in range(MAX_TRIPS - 1):
                trip_values.extend(
                    [
                        None,
                        None,
                        0,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                    ]
                )

            cur.execute(_INSERT_SQL, fixed_values + trip_values)
            inserted += 1

        conn.commit()
        log_db_op(
            f"استيراد تسجيل: {inserted} صف / {total_pax} راكب",
            operation="insert",
            query_summary="records",
            rows_affected=inserted,
            script_name="import_record",
            func_name="import_record",
        )

    except Exception as e:
        conn.rollback()
        log_exception(
            "فشل استيراد التسجيل",
            exc=e,
            script_name="import_record",
            func_name="import_record",
        )
        raise
    finally:
        conn.close()

    return {
        "rows": len(raw_rows),
        "total_pax": total_pax,
        "trips": inserted,
        "mapped_cols": list(col_map.keys()),
        "unmapped_cols": unmapped,
    }


# === التشغيل المباشر للاختبار ===
if __name__ == "__main__":
    print("هذا الملف يُستورد من السكربتات — للاختبار:")
    print("  from db.import_record import import_record")
    print('  stats = import_record("record.xlsx", date="01-01-2026", shift="A")')
