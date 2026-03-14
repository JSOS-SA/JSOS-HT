"""
مزامنة ملفات الإكسل من مجلد Record_2026 إلى قاعدة البيانات.

يراقب المجلد تلقائياً ويسجّل كل ملف جديد أو صف جديد
في الجداول: excel_folders, excel_files, records.

تشغيل مباشر: python db/excel_sync.py
"""

import os
import re
import sqlite3
import time
import logging
from pathlib import Path
from datetime import datetime

import openpyxl
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

from constants import MAX_TRIPS, TRIP_FIELDS
from helpers import build_records_insert_sql, safe_int, safe_str
from schema import get_connection, DB_PATH

# إعداد السجل
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("excel_sync")

# مسار مجلد السجلات
RECORD_DIR = Path(__file__).parent.parent / "monitor" / "Record_2026"

# عدد الأعمدة الثابتة قبل حقول الرحلات (التنسيق الجديد)
# رمز النوبة [0] + اسم المرسل [1] + وقت الرسالة [2]
FIXED_COLS_NEW = 3

# عدد الأعمدة الثابتة (التنسيق القديم)
# رمز النوبة [0] فقط — المرسل والوقت في آخر الصف
FIXED_COLS_OLD = 1


def parse_folder_info(folder_path: Path) -> dict | None:
    """استخراج معلومات الشهر واليوم من مسار المجلد."""
    name = folder_path.name
    # التحقق من صيغة اسم المجلد: DD-MM-YYYY
    if not re.match(r"^\d{2}-\d{2}-\d{4}$", name):
        return None
    parent_name = folder_path.parent.name
    # التحقق من صيغة مجلد الشهر: MM-YYYY
    if not re.match(r"^\d{2}-\d{4}$", parent_name):
        return None
    return {"month": parent_name, "day": name, "path": str(folder_path)}


def parse_file_info(file_path: Path) -> dict | None:
    """استخراج معلومات النوبة والنوع من اسم ملف الإكسل."""
    name = file_path.name
    if not name.endswith(".xlsx") or name.endswith(".tmp"):
        return None
    # صيغة الاسم: D-MM-YYYY-Record-X.xlsx أو DD-MM-YYYY-Record-X.xlsx
    # أو مع Stuck: D-MM-YYYY-Record-X-Stuck.xlsx
    match = re.match(
        r"^\d{1,2}-\d{2}-\d{4}-Record-([ABC])(-Stuck)?\.xlsx$", name
    )
    if not match:
        return None
    shift_code = match.group(1)
    is_stuck = match.group(2) is not None
    return {
        "filename": name,
        "shift_code": shift_code,
        "file_type": "stuck" if is_stuck else "normal",
    }


def get_or_create_sender(conn: sqlite3.Connection, name: str) -> int:
    """استرجاع معرّف المرسل أو إنشاؤه إذا لم يكن موجوداً."""
    if not name or not re.search(r"[\u0600-\u06FFa-zA-Z]{2,}", name):
        name = "غير معروف"
    row = conn.execute(
        "SELECT id FROM senders WHERE name = ?", (name,)
    ).fetchone()
    if row:
        return row[0]
    cur = conn.execute(
        "INSERT INTO senders (name) VALUES (?)", (name,)
    )
    return cur.lastrowid


def get_or_create_folder(conn: sqlite3.Connection, info: dict) -> int:
    """استرجاع معرّف المجلد أو إنشاؤه."""
    row = conn.execute(
        "SELECT id FROM excel_folders WHERE day = ?", (info["day"],)
    ).fetchone()
    if row:
        return row[0]
    cur = conn.execute(
        "INSERT INTO excel_folders (month, day, path) VALUES (?, ?, ?)",
        (info["month"], info["day"], info["path"]),
    )
    return cur.lastrowid


def get_or_create_file(
    conn: sqlite3.Connection, folder_id: int, info: dict
) -> int:
    """استرجاع معرّف الملف أو إنشاؤه."""
    row = conn.execute(
        "SELECT id FROM excel_files WHERE folder_id = ? AND filename = ?",
        (folder_id, info["filename"]),
    ).fetchone()
    if row:
        return row[0]
    cur = conn.execute(
        "INSERT INTO excel_files "
        "(folder_id, shift_code, filename, file_type) "
        "VALUES (?, ?, ?, ?)",
        (folder_id, info["shift_code"], info["filename"], info["file_type"]),
    )
    return cur.lastrowid


def _detect_format(ws) -> bool:
    """كشف تنسيق الملف — True إذا التنسيق الجديد (المرسل في العمود 2)."""
    header_2 = str(ws.cell(1, 2).value or "")
    return "المرسل" in header_2 or "اسم" in header_2


# جملة الإدراج — تُبنى مرة واحدة عند تحميل الوحدة
_INSERT_SQL = build_records_insert_sql()


def sync_excel_file(conn: sqlite3.Connection, file_path: Path) -> int:
    """مزامنة ملف إكسل واحد — إرجاع عدد الصفوف الجديدة."""
    folder_path = file_path.parent
    folder_info = parse_folder_info(folder_path)
    if not folder_info:
        return 0
    file_info = parse_file_info(file_path)
    if not file_info:
        return 0

    # تسجيل المجلد والملف
    folder_id = get_or_create_folder(conn, folder_info)
    file_id = get_or_create_file(conn, folder_id, file_info)

    # عدد الصفوف المسجّلة حالياً لهذا الملف
    existing_count = conn.execute(
        "SELECT COUNT(*) FROM records WHERE file_id = ?", (file_id,)
    ).fetchone()[0]

    # قراءة ملف الإكسل
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
    except Exception as e:
        logger.error(f"فشل قراءة الملف {file_path.name}: {e}")
        return 0

    ws = wb.active
    if ws is None or ws.max_row is None or ws.max_row < 2:
        wb.close()
        return 0

    total_rows = ws.max_row - 1  # بدون صف العناوين
    new_rows = 0

    if total_rows <= existing_count:
        wb.close()
        return 0

    col_count = ws.max_column
    is_new_format = _detect_format(ws)
    date_str = folder_info["day"]

    # قراءة الصفوف الجديدة فقط
    start_row = existing_count + 2  # +1 للعنوان +1 للصف التالي
    for row_idx in range(start_row, ws.max_row + 1):
        cells = []
        for col in range(1, col_count + 1):
            cells.append(ws.cell(row_idx, col).value)

        # تجاهل الصفوف الفارغة
        if not any(c is not None for c in cells):
            continue

        # استخراج المرسل والوقت حسب التنسيق
        if is_new_format:
            # الجديد: [0] نوبة، [1] مرسل، [2] وقت، [3+] رحلات
            sender_name = safe_str(cells[1]) if len(cells) > 1 else ""
            msg_time = safe_str(cells[2]) if len(cells) > 2 else ""
            trip_start = FIXED_COLS_NEW
        else:
            # القديم: [0] نوبة، [1+] رحلات، [آخر 2] وقت + مرسل
            last_idx = 0
            for i in range(len(cells) - 1, -1, -1):
                if cells[i] is not None:
                    last_idx = i
                    break
            effective_len = last_idx + 1
            row_trips_old = max(1, (effective_len - 3) // 9)
            msg_time_idx = FIXED_COLS_OLD + (row_trips_old * 9)
            sender_idx = msg_time_idx + 1
            msg_time = safe_str(cells[msg_time_idx]) if msg_time_idx < len(cells) else ""
            sender_name = safe_str(cells[sender_idx]) if sender_idx < len(cells) else ""
            trip_start = FIXED_COLS_OLD

        # حساب عدد الرحلات
        remaining = col_count - trip_start
        max_possible = remaining // 9
        trip_count = 0

        for t in range(min(max_possible, MAX_TRIPS)):
            base = trip_start + (t * 9)
            trip_slice = cells[base:base + 9]
            if any(c is not None for c in trip_slice):
                trip_count = t + 1

        if trip_count == 0:
            trip_count = 1

        # بناء قيم الرحلات (9 حقول × 10 رحلات = 90 قيمة)
        trip_values = []
        for t in range(MAX_TRIPS):
            base = trip_start + (t * 9)
            for i, field in enumerate(TRIP_FIELDS):
                idx = base + i
                if t < trip_count and idx < len(cells) and cells[idx] is not None:
                    if field == "passenger_count":
                        trip_values.append(safe_int(cells[idx]))
                    else:
                        trip_values.append(safe_str(cells[idx]))
                else:
                    trip_values.append(0 if field == "passenger_count" else None)

        # الأعمدة الثابتة
        fixed_values = [
            date_str,
            file_info["shift_code"],
            sender_name,
            msg_time,
            row_idx - 1,
            trip_count,
            file_id,
        ]

        conn.execute(_INSERT_SQL, fixed_values + trip_values)
        new_rows += 1

    # تحديث عدد الصفوف وتاريخ المزامنة
    conn.execute(
        "UPDATE excel_files SET row_count = ?, last_synced = ? "
        "WHERE id = ?",
        (total_rows, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), file_id),
    )
    conn.commit()
    wb.close()

    if new_rows > 0:
        logger.info(
            f"مزامنة {file_path.name}: {new_rows} صف جديد"
        )
    return new_rows


def full_sync(conn: sqlite3.Connection) -> int:
    """مزامنة كاملة لجميع ملفات الإكسل الموجودة."""
    if not RECORD_DIR.exists():
        logger.warning(f"المجلد غير موجود: {RECORD_DIR}")
        return 0

    total = 0
    for xlsx_file in sorted(RECORD_DIR.rglob("*.xlsx")):
        if xlsx_file.name.endswith(".tmp"):
            continue
        total += sync_excel_file(conn, xlsx_file)

    logger.info(f"المزامنة الكاملة: {total} صف جديد")
    return total


class ExcelFileHandler(FileSystemEventHandler):
    """مراقب أحداث الملفات — يزامن عند إنشاء أو تعديل ملف إكسل."""

    def __init__(self, db_path: str):
        """تهيئة المراقب مع مسار القاعدة."""
        super().__init__()
        self.db_path = db_path

    def _handle(self, event):
        """معالجة حدث ملف."""
        if event.is_directory:
            return
        file_path = Path(event.src_path)
        if not file_path.name.endswith(".xlsx"):
            return
        if file_path.name.endswith(".tmp"):
            return
        # انتظار قصير للتأكد من اكتمال الكتابة
        time.sleep(0.5)
        try:
            conn = get_connection(self.db_path)
            conn.execute("PRAGMA busy_timeout = 5000")
            sync_excel_file(conn, file_path)
            conn.close()
        except Exception as e:
            logger.error(f"خطأ مزامنة {file_path.name}: {e}")

    def on_created(self, event):
        """عند إنشاء ملف جديد."""
        self._handle(event)

    def on_modified(self, event):
        """عند تعديل ملف موجود."""
        self._handle(event)

    def on_moved(self, event):
        """عند إعادة تسمية ملف (النمط المؤقت ثم الاستبدال)."""
        if hasattr(event, "dest_path"):
            dest = Path(event.dest_path)
            if dest.name.endswith(".xlsx") and not dest.name.endswith(
                ".tmp"
            ):
                try:
                    conn = get_connection(self.db_path)
                    conn.execute("PRAGMA busy_timeout = 5000")
                    sync_excel_file(conn, dest)
                    conn.close()
                except Exception as e:
                    logger.error(f"خطأ مزامنة {dest.name}: {e}")


def start_watcher(db_path: str = None):
    """تشغيل المراقب التلقائي."""
    db = db_path or DB_PATH
    if not RECORD_DIR.exists():
        logger.error(f"المجلد غير موجود: {RECORD_DIR}")
        return

    # مزامنة كاملة أولاً
    conn = get_connection(db)
    conn.execute("PRAGMA busy_timeout = 5000")
    full_sync(conn)
    conn.close()

    # تشغيل المراقب
    handler = ExcelFileHandler(db)
    observer = Observer()
    observer.schedule(handler, str(RECORD_DIR), recursive=True)
    observer.start()
    logger.info(f"المراقب يعمل على: {RECORD_DIR}")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
        logger.info("تم إيقاف المراقب")
    observer.join()


# === التشغيل المباشر ===
if __name__ == "__main__":
    start_watcher()
