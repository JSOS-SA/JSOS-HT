"""
كتابة سجلات القاعدة الجديدة في ملفات الإكسل.

يقرأ من record_2026 السجلات التي لم تُكتب بعد (filename IS NULL)
ويكتبها في ملفات الإكسل بنفس الهيكل المعتمد، ثم يحدّث القاعدة.

تشغيل مرة واحدة: python db/write_excel_from_db.py
تشغيل دوري: python db/write_excel_from_db.py --watch

TODO: المزامنة التلقائية من القاعدة إلى الإكسل متوقفة حالياً.
- لا يُنشئ ملف إكسل جديد لكل نوبة تلقائياً.
- البيانات الأخيرة في القاعدة ليس لها ملفات إكسل.
- المفترض أن المزامنة تعمل تلقائياً عبر --watch.
- يُعالج كمهمة منفصلة.
"""

import argparse
import logging
import sqlite3
import time
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from constants import MAX_TRIPS, TRIP_FIELD_NAMES_AR, TRIP_FIELDS
from helpers import safe_str

# ═══════════════════════════════════════════════════════
# الإعدادات
# ═══════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("write_excel_from_db")

# مسارات
DB_DIR = Path(__file__).parent
DB_PATH = DB_DIR / "ht_sc_new.db"
RECORD_DIR = Path(__file__).parent.parent / "Record_2026"

# الفاصل الزمني للمراقبة الدورية (ثوانٍ)
WATCH_INTERVAL = 5

# اسم مستعار للتوافق — نفس قيم TRIP_FIELDS بالضبط
TRIP_DB_FIELDS = TRIP_FIELDS

# عدد الأعمدة الثابتة: م + رمز النوبة + اسم المرسل + وقت وصول الرسالة
FIXED_COLS = 4


# ═══════════════════════════════════════════════════════
# 1. اتصال القاعدة
# ═══════════════════════════════════════════════════════


def get_connection(db_path: str | Path | None = None) -> sqlite3.Connection:
    """إنشاء اتصال بالقاعدة الجديدة مع تفعيل المفاتيح الأجنبية."""
    path = str(db_path or DB_PATH)
    conn = sqlite3.connect(path)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    return conn


def migrate_filename_nullable(conn: sqlite3.Connection) -> None:
    """ترحيل عمود filename ليقبل NULL — مطلوب لإدراج المراقب مباشرة."""
    cursor = conn.execute("PRAGMA table_info(record_2026)")
    columns = cursor.fetchall()
    filename_col = None
    for col in columns:
        if col["name"] == "filename":
            filename_col = col
            break

    if filename_col is None or filename_col["notnull"] == 0:
        return  # يقبل NULL بالفعل أو العمود غير موجود

    # استخراج تعريف الجدول الأصلي
    row = conn.execute(
        "SELECT sql FROM sqlite_master "
        "WHERE type='table' AND name='record_2026'"
    ).fetchone()
    if not row:
        return

    # تعديل التعريف: إزالة NOT NULL من filename وتسمية مؤقتة
    import re

    new_sql = re.sub(
        r"filename\s+TEXT\s+NOT\s+NULL",
        "filename        TEXT",
        row["sql"],
        flags=re.IGNORECASE,
    )
    new_sql = new_sql.replace(
        "CREATE TABLE record_2026",
        "CREATE TABLE record_2026_tmp",
        1,
    )

    conn.execute("BEGIN IMMEDIATE")
    try:
        conn.executescript(f"""
            {new_sql};
            INSERT INTO record_2026_tmp SELECT * FROM record_2026;
            DROP TABLE record_2026;
        """)
        conn.execute(
            "ALTER TABLE record_2026_tmp RENAME TO record_2026"
        )
        # إعادة إنشاء الفهارس
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_record2026_date "
            "ON record_2026(folder_date)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_record2026_shift "
            "ON record_2026(shift_code)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_record2026_flight "
            "ON record_2026(flight_number_1)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_record2026_sender "
            "ON record_2026(sender_name)"
        )
        conn.commit()
        logger.info("تم ترحيل عمود filename ليقبل NULL")
    except Exception:
        conn.rollback()
        raise


# ═══════════════════════════════════════════════════════
# 2. أدوات التحويل
# ═══════════════════════════════════════════════════════


def _shift_full_code(letter: str) -> str:
    """تحويل حرف النوبة لرمز كامل: A → 00A#."""
    return f"00{letter}#"


# ═══════════════════════════════════════════════════════
# 3. بناء الإكسل
# ═══════════════════════════════════════════════════════


def _build_excel_path(folder_date: str, shift_code: str) -> Path:
    """بناء مسار ملف الإكسل من التاريخ وحرف النوبة.

    المسار: Record_2026/MM-YYYY/DD-MM-YYYY/DD-MM-YYYY-Record-X.xlsx
    """
    parts = folder_date.split("-")
    if len(parts) != 3:
        msg = f"صيغة تاريخ غير صالحة: {folder_date}"
        raise ValueError(msg)

    # DD-MM-YYYY → MM-YYYY للمجلد الشهري
    month_dir = f"{parts[1]}-{parts[2]}"
    # حشو اليوم بالصفر إذا لم يكن محشواً
    day_padded = parts[0].zfill(2)
    day_dir = f"{day_padded}-{parts[1]}-{parts[2]}"
    filename = f"{day_dir}-Record-{shift_code}.xlsx"

    return RECORD_DIR / month_dir / day_dir / filename


def _build_header_row(max_trips: int) -> list[str]:
    """بناء صف العناوين حسب عدد الرحلات.

    رحلة واحدة: بدون لاحقة. أكثر: لاحقة رقمية.
    """
    header: list[str] = ["م", "رمز النوبة", "اسم المرسل", "وقت وصول الرسالة"]
    for t in range(max_trips):
        suffix = f" {t + 1}" if max_trips > 1 else ""
        for name in TRIP_FIELD_NAMES_AR:
            header.append(name + suffix)
    return header


def _build_data_row(
    record: sqlite3.Row,
    max_trips: int,
) -> list[object]:
    """بناء صف البيانات من سجل القاعدة.

    الترتيب: [id, رمز النوبة الكامل, المرسل, الوقت, ...حقول الرحلات]
    """
    trip_count = record["trip_count"] or 1
    row: list[object] = [
        record["id"],
        _shift_full_code(record["shift_code"]),
        safe_str(record["sender_name"]),
        safe_str(record["msg_time"]),
    ]

    for t in range(1, max_trips + 1):
        for field in TRIP_DB_FIELDS:
            col_name = f"{field}_{t}"
            val = record[col_name] if t <= trip_count else None
            if field == "passenger_count":
                row.append(val if val else 0)
            else:
                row.append(safe_str(val))

    return row


# ═══════════════════════════════════════════════════════
# 4. كتابة الإكسل
# ═══════════════════════════════════════════════════════


def _write_batch_to_excel(
    records: list[sqlite3.Row],
    folder_date: str,
    shift_code: str,
) -> list[tuple[int, str, int]]:
    """كتابة مجموعة سجلات في ملف إكسل واحد.

    إرجاع: قائمة (record_id, filename, row_num) لتحديث القاعدة.
    """
    file_path = _build_excel_path(folder_date, shift_code)

    # إنشاء المجلد إذا لم يكن موجوداً
    file_path.parent.mkdir(parents=True, exist_ok=True)

    # حساب أقصى عدد رحلات بين السجلات الجديدة
    new_max_trips = max((r["trip_count"] or 1) for r in records)

    # قراءة الملف الموجود أو إنشاء جديد
    wb = openpyxl.Workbook()
    current_max_trips = 1
    has_id_column = False
    file_existed = False

    if file_path.exists():
        file_existed = True
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            if ws is not None and ws.max_row and ws.max_row >= 1:
                first_header = safe_str(ws.cell(1, 1).value)
                has_id_column = first_header == "م"
                header_len = ws.max_column or 0
                fixed = FIXED_COLS if has_id_column else 3
                current_max_trips = max(
                    1, (header_len - fixed) // 9
                )
        except Exception:
            # ملف تالف — إنشاء جديد
            wb = openpyxl.Workbook()
            file_existed = False

    # الحصول على ورقة العمل النشطة
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("رسائل")

    # ترحيل ملف قديم بدون عمود المعرّف — فقط للملفات الموجودة مسبقاً
    if file_existed and not has_id_column and ws.max_row and ws.max_row > 0:
        for r in range(ws.max_row, 0, -1):
            cell_count = ws.max_column or 0
            for c in range(cell_count, 0, -1):
                ws.cell(r, c + 1).value = ws.cell(r, c).value  # type: ignore[assignment]
            ws.cell(r, 1).value = "م" if r == 1 else ""  # type: ignore[assignment]
        has_id_column = True

    final_max_trips = max(current_max_trips, new_max_trips)

    # كتابة العناوين — صراحةً بدون الاعتماد على max_row
    if not file_existed:
        # ملف جديد — كتابة العناوين في الصف الأول
        header = _build_header_row(final_max_trips)
        for i, val in enumerate(header, 1):
            ws.cell(1, i).value = val  # type: ignore[assignment]
        for i in range(1, len(header) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18
    elif final_max_trips > current_max_trips:
        # توسيع العناوين لملف موجود
        header = _build_header_row(final_max_trips)
        for i, val in enumerate(header, 1):
            ws.cell(1, i).value = val  # type: ignore[assignment]
        for i in range(1, len(header) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18

    # رقم الصف التالي — للملف الجديد: 2 (بعد العناوين)
    if not file_existed:
        next_row = 2
    else:
        next_row = (ws.max_row or 1) + 1

    # كتابة صفوف البيانات — صراحةً بدون ws.append
    updates: list[tuple[int, str, int]] = []
    for i, record in enumerate(records):
        data_row = _build_data_row(record, final_max_trips)
        row_num = next_row + i
        for j, val in enumerate(data_row, 1):
            ws.cell(row_num, j).value = val  # type: ignore[assignment]
        updates.append((record["id"], file_path.name, row_num))

    # كتابة آمنة — ملف مؤقت ثم استبدال
    tmp_path = file_path.with_suffix(".xlsx.tmp")
    wb.save(str(tmp_path))
    tmp_path.replace(file_path)

    return updates


# ═══════════════════════════════════════════════════════
# 5. المزامنة
# ═══════════════════════════════════════════════════════


def _auto_repair(conn: sqlite3.Connection) -> int:
    """فحص ملفات الإكسل المعطوبة وإعادة تعيينها للكتابة من جديد."""
    rows = conn.execute(
        "SELECT DISTINCT filename, folder_date, shift_code "
        "FROM record_2026 WHERE filename IS NOT NULL"
    ).fetchall()

    repaired = 0
    for row in rows:
        fname = row["filename"]
        fdate = row["folder_date"]
        scode = row["shift_code"]

        try:
            fpath = _build_excel_path(fdate, scode)
        except ValueError:
            continue

        # ملف مفقود أو عناوين معطوبة — إعادة تعيين
        needs_repair = False
        if not fpath.exists():
            needs_repair = True
        else:
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True)
                ws = wb.active
                if ws is None:
                    needs_repair = True
                else:
                    second_cell = safe_str(ws.cell(1, 2).value)
                    if second_cell != "رمز النوبة":
                        needs_repair = True
                wb.close()
            except Exception:
                needs_repair = True

        if needs_repair:
            # حذف الملف المعطوب
            if fpath.exists():
                fpath.unlink()
            # إعادة تعيين السجلات للكتابة من جديد
            conn.execute(
                "UPDATE record_2026 SET filename = NULL, row_num = NULL "
                "WHERE filename = ?",
                (fname,),
            )
            repaired += 1

    if repaired:
        conn.commit()
        logger.info(f"إصلاح تلقائي: {repaired} ملف أُعيد تعيينه")

    return repaired


def get_pending_records(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    """قراءة السجلات التي لم تُكتب في الإكسل بعد."""
    return conn.execute(
        "SELECT * FROM record_2026 "
        "WHERE filename IS NULL "
        "ORDER BY id"
    ).fetchall()


def update_records(
    conn: sqlite3.Connection,
    updates: list[tuple[int, str, int]],
) -> None:
    """تحديث filename و row_num بعد كتابة الإكسل."""
    conn.executemany(
        "UPDATE record_2026 "
        "SET filename = ?, row_num = ? "
        "WHERE id = ?",
        [(fname, rnum, rid) for rid, fname, rnum in updates],
    )
    conn.commit()


def refresh_shift_stats(conn: sqlite3.Connection) -> None:
    """حساب إحصائيات كل نوبة من record_2026 وتعبئة جداول النوبات."""
    for s in ("A", "B", "C"):
        table = f"shift_{s.lower()}_stats"

        # شروط الفيزا
        umrah_visa = " OR ".join(
            f"visa_type_{n} LIKE '%عمر%'"
            for n in range(1, MAX_TRIPS + 1)
        )
        tourist_visa = " OR ".join(
            f"visa_type_{n} LIKE '%سياح%'"
            for n in range(1, MAX_TRIPS + 1)
        )

        # مجموع الركاب
        total_pax = " + ".join(
            f"COALESCE(passenger_count_{n}, 0)"
            for n in range(1, MAX_TRIPS + 1)
        )
        umrah_pax = " + ".join(
            f"CASE WHEN visa_type_{n} LIKE '%عمر%'"
            f" THEN COALESCE(passenger_count_{n}, 0) ELSE 0 END"
            for n in range(1, MAX_TRIPS + 1)
        )
        tourist_pax = " + ".join(
            f"CASE WHEN visa_type_{n} LIKE '%سياح%'"
            f" THEN COALESCE(passenger_count_{n}, 0) ELSE 0 END"
            for n in range(1, MAX_TRIPS + 1)
        )

        # التفويج
        early = " + ".join(
            f"CASE WHEN dispatch_{n} LIKE '%مبكر%'"
            f" THEN 1 ELSE 0 END"
            for n in range(1, MAX_TRIPS + 1)
        )
        late = " + ".join(
            f"CASE WHEN dispatch_{n} LIKE '%تأخر%'"
            f" OR dispatch_{n} LIKE '%تاخر%'"
            f" THEN 1 ELSE 0 END"
            for n in range(1, MAX_TRIPS + 1)
        )
        wrong = " + ".join(
            f"CASE WHEN dispatch_{n} LIKE '%خاط%'"
            f" THEN 1 ELSE 0 END"
            for n in range(1, MAX_TRIPS + 1)
        )
        mix = " + ".join(
            f"CASE WHEN dispatch_{n} LIKE '%مشترك%'"
            f" THEN 1 ELSE 0 END"
            for n in range(1, MAX_TRIPS + 1)
        )

        conn.execute(f"DELETE FROM {table}")
        conn.execute(f"""
            INSERT INTO {table} (
                folder_date, dep_buses, umrah_buses, tourist_buses,
                dep_pax, umrah_pax, tourist_pax,
                arrival_buses, bus_accidents, car_accidents,
                early, late, wrong, mix, minutes,
                stumbler,
                operational_letters, weather_conditions,
                operational_reasons, emergency_maintenance,
                other_notes
            )
            SELECT
                folder_date,
                COUNT(*),
                SUM(CASE WHEN {umrah_visa}
                    THEN 1 ELSE 0 END),
                SUM(CASE WHEN {tourist_visa}
                    THEN 1 ELSE 0 END),
                SUM({total_pax}),
                SUM({umrah_pax}),
                SUM({tourist_pax}),
                0, 0, 0,
                SUM({early}),
                SUM({late}),
                SUM({wrong}),
                SUM({mix}),
                SUM({early}) + SUM({late})
                    + SUM({wrong}) + SUM({mix}),
                NULL,
                0, 0, 0, 0, 0
            FROM record_2026
            WHERE shift_code = '{s}'
            GROUP BY folder_date
        """)

    conn.commit()


def refresh_daily_stats(conn: sqlite3.Connection) -> None:
    """تعبئة الجدول اليومي بجمع النوبات الثلاث لكل يوم."""
    conn.execute("DELETE FROM daily_stats")
    conn.execute("""
        INSERT INTO daily_stats (
            folder_date, dep_buses, umrah_buses, tourist_buses,
            dep_pax, umrah_pax, tourist_pax,
            arrival_buses, bus_accidents, car_accidents,
            early, late, wrong, mix,
            stumbler,
            operational_letters, weather_conditions,
            operational_reasons, emergency_maintenance,
            other_notes, total
        )
        SELECT
            folder_date,
            SUM(dep_buses),
            SUM(umrah_buses),
            SUM(tourist_buses),
            SUM(dep_pax),
            SUM(umrah_pax),
            SUM(tourist_pax),
            SUM(arrival_buses),
            SUM(bus_accidents),
            SUM(car_accidents),
            SUM(early),
            SUM(late),
            SUM(wrong),
            SUM(mix),
            NULL,
            SUM(operational_letters),
            SUM(weather_conditions),
            SUM(operational_reasons),
            SUM(emergency_maintenance),
            SUM(other_notes),
            SUM(dep_buses) + SUM(umrah_buses) + SUM(tourist_buses)
            + SUM(dep_pax) + SUM(umrah_pax) + SUM(tourist_pax)
            + SUM(arrival_buses) + SUM(bus_accidents) + SUM(car_accidents)
            + SUM(early) + SUM(late) + SUM(wrong) + SUM(mix)
        FROM (
            SELECT * FROM shift_a_stats
            UNION ALL
            SELECT * FROM shift_b_stats
            UNION ALL
            SELECT * FROM shift_c_stats
        )
        GROUP BY folder_date
    """)
    conn.commit()


def sync_once(conn: sqlite3.Connection) -> int:
    """دورة مزامنة واحدة — إصلاح + قراءة السجلات المعلّقة + كتابتها."""
    # إصلاح تلقائي للملفات المعطوبة قبل الكتابة
    _auto_repair(conn)

    records = get_pending_records(conn)
    if not records:
        return 0

    # تجميع السجلات حسب (التاريخ, حرف النوبة)
    groups: dict[tuple[str, str], list[sqlite3.Row]] = defaultdict(list)
    for rec in records:
        key = (rec["folder_date"], rec["shift_code"])
        groups[key].append(rec)

    total = 0
    all_updates: list[tuple[int, str, int]] = []

    for (folder_date, shift_code), batch in groups.items():
        try:
            updates = _write_batch_to_excel(batch, folder_date, shift_code)
            all_updates.extend(updates)
            total += len(updates)
            logger.info(
                f"{folder_date}/{shift_code}: "
                f"{len(updates)} صف في الإكسل"
            )
        except Exception as e:
            logger.error(
                f"خطأ كتابة {folder_date}/{shift_code}: {e}"
            )

    if all_updates:
        update_records(conn, all_updates)

    # تحديث جداول الإحصائيات الأربعة بعد كل دورة ناجحة
    refresh_shift_stats(conn)
    refresh_daily_stats(conn)

    return total


# ═══════════════════════════════════════════════════════
# 6. التشغيل
# ═══════════════════════════════════════════════════════


def main() -> None:
    """نقطة الدخول — مزامنة مرة أو مراقبة دورية."""
    parser = argparse.ArgumentParser(
        description="كتابة سجلات القاعدة في ملفات الإكسل"
    )
    parser.add_argument(
        "--watch",
        action="store_true",
        help="تشغيل دوري كل 5 ثوانٍ",
    )
    args = parser.parse_args()

    conn = get_connection()
    migrate_filename_nullable(conn)

    if args.watch:
        logger.info(
            f"المراقبة الدورية كل {WATCH_INTERVAL} ثوانٍ..."
        )
        try:
            while True:
                written = sync_once(conn)
                if written > 0:
                    logger.info(f"تمت كتابة {written} صف")
                time.sleep(WATCH_INTERVAL)
        except KeyboardInterrupt:
            logger.info("تم الإيقاف")
        finally:
            conn.close()
    else:
        written = sync_once(conn)
        logger.info(f"المزامنة: {written} صف جديد")
        conn.close()


if __name__ == "__main__":
    main()
